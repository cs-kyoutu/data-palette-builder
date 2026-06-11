"""BI 専用 Excel 出力 — シート①ウィザード手順書 / シート②生成SQLプレビュー。

データパレットの excel_builder とは別物(テンプレートが違う)。出力先は backend/output。
"""
from __future__ import annotations

import uuid
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

_OUTPUT_DIR = Path(__file__).parent.parent / "output"
_OUTPUT_DIR.mkdir(exist_ok=True)


def build_bi_spreadsheet(design_doc: dict, procedure_text: str, sql: str) -> tuple[str, str]:
    """手順書テキスト + SQL から BI 用 Excel を生成。 (filepath, filename) を返す。"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "手順書"
    ws1.column_dimensions["A"].width = 100
    title = Font(bold=True, size=12)
    for i, line in enumerate(procedure_text.split("\n"), start=1):
        cell = ws1.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("【") or line.startswith("■"):
            cell.font = title

    ws2 = wb.create_sheet("生成SQL")
    ws2.column_dimensions["A"].width = 100
    ws2.cell(row=1, column=1, value="※ 参考用の等価SQLです。本ツールはこのSQLを実行しません。").font = Font(italic=True)
    for i, line in enumerate(sql.split("\n"), start=2):
        ws2.cell(row=i, column=1, value=line).alignment = Alignment(vertical="top")

    rtype = design_doc.get("report_type", "report")
    filename = f"bi_{rtype}_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = _OUTPUT_DIR / filename
    wb.save(filepath)
    return str(filepath), filename


def _sheet_safe(name: str, used: set) -> str:
    """Excel シート名は31文字以内・一部記号不可・重複不可。安全名に変換する。"""
    base = "".join(ch for ch in str(name) if ch not in '[]:*?/\\')[:28] or "シート"
    cand, i = base, 2
    while cand in used:
        cand = f"{base[:26]}_{i}"
        i += 1
    used.add(cand)
    return cand


def build_design_spreadsheet(design: dict, design_text: str) -> tuple[str, str]:
    """逆算設計の設計書テキスト + テーブル定義表から Excel を生成。 (filepath, filename)。

    シート①設計書(全文) / 各テーブルごとに テーブル定義(カラム表)・サンプルデータ。
    単一テーブルと複数テーブル(design に "テーブル":[...])の両方に対応。
    """
    wb = Workbook()
    title = Font(bold=True, size=12)
    header = Font(bold=True, color="FFFFFF")
    used_names = {"設計書"}

    # シート① 設計書(全文)
    ws1 = wb.active
    ws1.title = "設計書"
    ws1.column_dimensions["A"].width = 100
    for i, line in enumerate(design_text.split("\n"), start=1):
        cell = ws1.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("【") or line.startswith("■") or line.startswith("▼"):
            cell.font = title

    # 複数テーブル構成なら各テーブル、単一なら design 自身を1テーブル扱い
    _t = design.get("テーブル")
    tables = _t if isinstance(_t, list) and _t else [design]
    multi = len(tables) > 1 or (isinstance(_t, list) and _t)

    for idx, t in enumerate(tables, 1):
        tname = t.get("テーブル名") or (f"テーブル{idx}" if multi else "")
        label = tname or "テーブル定義"

        # テーブル定義シート
        ws = wb.create_sheet(_sheet_safe(("定義_" + tname) if multi else "テーブル定義", used_names))
        if multi:
            ws.append([f"テーブル: {tname} / 粒度: {t.get('粒度', '')} / 主キー: {t.get('主キー', '')}"])
            ws["A1"].font = title
            ws.append([])
        ws.append(["カラム名", "データ型", "主キー", "派生", "派生方法", "用途"])
        for c in ws[ws.max_row]:
            c.font = header
        for col in t.get("カラム", []) or []:
            ws.append([
                col.get("name", ""), col.get("type", ""),
                "○" if col.get("主キー") else "",
                "○" if col.get("derived") else "",
                col.get("派生方法", ""), str(col.get("用途", "")),
            ])
        for letter, width in zip("ABCDEF", (24, 12, 8, 8, 36, 40)):
            ws.column_dimensions[letter].width = width

        # サンプルデータシート
        sample = t.get("サンプル") or {}
        cols_order = sample.get("カラム順") or [c.get("name", "") for c in t.get("カラム", []) or []]
        rows = sample.get("行") or []
        if cols_order and rows:
            wss = wb.create_sheet(_sheet_safe(("試料_" + tname) if multi else "サンプルデータ", used_names))
            wss.append(list(cols_order))
            for c in wss[1]:
                c.font = header
            for row in rows:
                wss.append(list(row))

    filename = f"design_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = _OUTPUT_DIR / filename
    wb.save(filepath)
    return str(filepath), filename
