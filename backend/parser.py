"""
テーブル定義書パーサー
- インプット: テーブル定義書（Excel/CSV）→ テーブル名 + カラム定義リスト
- アウトプット: マッピングファイル（Excel/CSV）→ カラム名 + 定義 + ソース情報
"""
import csv
from pathlib import Path

import openpyxl


# --- ヘッダー行の自動検出用キーワード ---
INPUT_HEADER_KEYWORDS = {
    "カラム名", "項目名", "name", "column", "フィールド名", "フィールド",
    "カラム論理名", "カラム物理名", "列名", "属性名", "項目",
    "テーブル名", "データファイル名",
}
OUTPUT_HEADER_KEYWORDS = {
    "カラム名", "項目名", "name", "column", "列名", "フィールド名",
    "定義", "definition", "説明",
    "素材になる", "インプットカラム", "インプットデータ",
    "ソースカラム", "ソーステーブル", "source",
}

# b→dashメタデータ形式の検出キーワード
BDASH_META_KEYWORDS = {"データファイル名", "カラム論理名", "データ型"}


def _detect_header_row(ws, keywords: set[str]) -> int | None:
    """ヘッダー行を自動検出（キーワードに一致するセルがある行）"""
    keywords_lower = {k.lower() for k in keywords}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), start=1):
        for cell_val in row:
            if not cell_val:
                continue
            cell_str = str(cell_val).strip().lower().replace("\n", "")
            # 完全一致
            if cell_str in keywords_lower:
                return row_idx
            # 部分一致（改行入りヘッダー対応: "素材になる\nカラム名" 等）
            for kw in keywords_lower:
                if kw in cell_str or cell_str in kw:
                    return row_idx
    # 見つからなければ1行目をヘッダーとする
    return 1


def _get_col_index(header_row: list, *candidates: str) -> int | None:
    """ヘッダー行から指定キーワードに一致する列インデックスを返す（部分一致対応）"""
    candidates_lower = {c.lower() for c in candidates}
    for i, val in enumerate(header_row):
        if not val:
            continue
        val_str = str(val).strip().lower().replace("\n", "")
        # 完全一致
        if val_str in candidates_lower:
            return i
        # 部分一致
        for c in candidates_lower:
            if c in val_str or val_str in c:
                return i
    return None


# =============================================================
# b→dashメタデータ形式の判定・解析
# =============================================================

def _is_bdash_meta_format(header: list[str]) -> bool:
    """b→dashメタデータ形式かどうか判定"""
    header_lower = {str(v).strip().lower() for v in header if v}
    return len(BDASH_META_KEYWORDS & {h for h in header_lower if h in {"データファイル名", "カラム論理名", "データ型"}}) >= 2


def _parse_bdash_meta(rows: list, header: list[str]) -> list[dict]:
    """
    b→dashメタデータ形式を解析
    データファイル名でグループ化し、カラム論理名・データ型・顧客IDフラグを抽出
    返り値: [{table_name, columns: [{name, type, description, is_customer_id}]}]
    """
    tbl_col = _get_col_index(header, "データファイル名")
    name_col = _get_col_index(header, "カラム論理名")
    type_col = _get_col_index(header, "データ型")
    cid_col = _get_col_index(header, "顧客IDフラグ", "顧客idフラグ")

    if tbl_col is None or name_col is None:
        return []

    # データファイル名でグループ化
    tables_dict: dict[str, list[dict]] = {}
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        tbl_name = str(row[tbl_col]).strip() if len(row) > tbl_col and row[tbl_col] else ""
        col_name = str(row[name_col]).strip() if len(row) > name_col and row[name_col] else ""
        if not tbl_name or not col_name:
            continue

        col_type = str(row[type_col]).strip() if type_col is not None and len(row) > type_col and row[type_col] else ""
        is_cid = False
        if cid_col is not None and len(row) > cid_col and row[cid_col]:
            cid_val = str(row[cid_col]).strip().lower()
            is_cid = cid_val in {"true", "1", "○", "yes"}

        if tbl_name not in tables_dict:
            tables_dict[tbl_name] = []
        col_entry = {"name": col_name, "type": col_type, "description": ""}
        if is_cid:
            col_entry["description"] = "顧客ID（結合キー）"
        tables_dict[tbl_name].append(col_entry)

    return [{"table_name": name, "columns": cols} for name, cols in tables_dict.items()]


# =============================================================
# インプットテーブル定義書の解析
# =============================================================

def parse_input_excel(file_path: str) -> list[dict]:
    """
    インプットテーブル定義書（Excel）を解析
    返り値: [{table_name, columns: [{name, type, description}]}]
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    tables = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 2:
            continue

        header_row_idx = _detect_header_row(ws, INPUT_HEADER_KEYWORDS)
        rows = list(ws.iter_rows(min_row=header_row_idx, values_only=True))
        if not rows:
            continue

        header = [str(v).strip() if v else "" for v in rows[0]]

        # b→dashメタデータ形式の自動検出
        if _is_bdash_meta_format(header):
            meta_tables = _parse_bdash_meta(rows[1:], header)
            tables.extend(meta_tables)
            continue

        # 通常形式: カラム名、型、説明の列を特定
        name_col = _get_col_index(header, "カラム名", "項目名", "name", "column", "フィールド名", "カラム論理名", "列名", "属性名", "項目", "カラム物理名")
        type_col = _get_col_index(header, "型", "type", "データ型", "形式", "data_type", "タイプ")
        desc_col = _get_col_index(header, "説明", "description", "備考", "定義", "概要", "用途", "コメント", "memo")

        if name_col is None:
            name_col = 0

        columns = []
        for row in rows[1:]:
            if not row or all(v is None for v in row):
                continue
            name_val = str(row[name_col]).strip() if row[name_col] else ""
            if not name_val:
                continue
            columns.append({
                "name": name_val,
                "type": str(row[type_col]).strip() if type_col is not None and row[type_col] else "",
                "description": str(row[desc_col]).strip() if desc_col is not None and row[desc_col] else "",
            })

        if columns:
            tables.append({"table_name": sheet_name, "columns": columns})

    wb.close()
    return tables


def parse_input_csv(file_path: str, table_name: str | None = None) -> list[dict]:
    """
    インプットテーブル定義書（CSV）を解析
    返り値: [{table_name, columns: [{name, type, description}]}]
    """
    if not table_name:
        table_name = Path(file_path).stem

    with open(file_path, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if len(rows) < 2:
        return []

    header = [v.strip() for v in rows[0]]

    # b→dashメタデータ形式の自動検出
    if _is_bdash_meta_format(header):
        return _parse_bdash_meta(rows[1:], header)

    name_col = _get_col_index(header, "カラム名", "項目名", "name", "column", "フィールド名", "カラム論理名")
    type_col = _get_col_index(header, "型", "type", "データ型", "形式", "data_type", "タイプ")
    desc_col = _get_col_index(header, "説明", "description", "備考", "定義", "概要", "用途", "コメント", "memo")

    if name_col is None:
        name_col = 0

    columns = []
    for row in rows[1:]:
        if not row:
            continue
        name_val = row[name_col].strip() if len(row) > name_col else ""
        if not name_val:
            continue
        columns.append({
            "name": name_val,
            "type": row[type_col].strip() if type_col is not None and len(row) > type_col else "",
            "description": row[desc_col].strip() if desc_col is not None and len(row) > desc_col else "",
        })

    return [{"table_name": table_name, "columns": columns}] if columns else []


# =============================================================
# アウトプットマッピングファイルの解析
# =============================================================

def parse_output_excel(file_path: str) -> dict:
    """
    アウトプットマッピングファイル（Excel）を解析
    フォーマット: カラム名 | 定義 | インプットカラム | インプットデータ
    返り値: {columns: [{name, definition, source_column, source_table}]}
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  # 最初のシートを使用

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return {"columns": []}

    # ヘッダー行を検出
    header_idx = 0
    header_keywords = ["カラム名", "項目名", "name", "column", "列名", "フィールド名", "定義", "素材になる"]
    for i, row in enumerate(rows):
        vals = [str(v).strip().lower().replace("\n", "") if v else "" for v in row]
        if any(k in v for v in vals for k in header_keywords):
            header_idx = i
            break

    header = [str(v).strip() if v else "" for v in rows[header_idx]]

    name_col = _get_col_index(header, "カラム名", "項目名", "name", "column", "列名", "フィールド名")
    def_col = _get_col_index(header, "定義", "説明", "definition", "description", "備考", "用途")
    src_col_col = _get_col_index(header, "インプットカラム", "元カラム", "source_column", "ソースカラム", "素材になるカラム名", "素材になる\nカラム名")
    src_tbl_col = _get_col_index(header, "インプットデータ", "元テーブル", "source_table", "ソーステーブル", "素材になるID/データファイル名", "素材になる\nID/データファイル名")

    if name_col is None:
        # フォールバック: 列順序で推定（B列=カラム名, C列=定義, D列=インプットカラム, E列=インプットデータ）
        name_col = 1
        def_col = 2
        src_col_col = 3
        src_tbl_col = 4

    columns = []
    for row in rows[header_idx + 1:]:
        if not row or all(v is None for v in row):
            continue
        name_val = str(row[name_col]).strip() if len(row) > name_col and row[name_col] else ""
        if not name_val:
            continue
        columns.append({
            "name": name_val,
            "definition": str(row[def_col]).strip() if def_col is not None and len(row) > def_col and row[def_col] else "",
            "source_column": str(row[src_col_col]).strip() if src_col_col is not None and len(row) > src_col_col and row[src_col_col] else "",
            "source_table": str(row[src_tbl_col]).strip() if src_tbl_col is not None and len(row) > src_tbl_col and row[src_tbl_col] else "",
        })

    return {"columns": columns}


def parse_output_csv(file_path: str) -> dict:
    """
    アウトプットマッピングファイル（CSV）を解析
    返り値: {columns: [{name, definition, source_column, source_table}]}
    """
    with open(file_path, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if len(rows) < 2:
        return {"columns": []}

    header = [v.strip() for v in rows[0]]
    name_col = _get_col_index(header, "カラム名", "項目名", "name")
    def_col = _get_col_index(header, "定義", "説明", "definition")
    src_col_col = _get_col_index(header, "インプットカラム", "元カラム", "source_column")
    src_tbl_col = _get_col_index(header, "インプットデータ", "元テーブル", "source_table")

    if name_col is None:
        name_col = 0

    columns = []
    for row in rows[1:]:
        if not row:
            continue
        name_val = row[name_col].strip() if len(row) > name_col else ""
        if not name_val:
            continue
        columns.append({
            "name": name_val,
            "definition": row[def_col].strip() if def_col is not None and len(row) > def_col else "",
            "source_column": row[src_col_col].strip() if src_col_col is not None and len(row) > src_col_col else "",
            "source_table": row[src_tbl_col].strip() if src_tbl_col is not None and len(row) > src_tbl_col else "",
        })

    return {"columns": columns}
