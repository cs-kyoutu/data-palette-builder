"""
データパレット構築手順書ジェネレータ
FastAPI + Claude API による自動手順書生成Webアプリ

フロー:
1. インプット: テーブル定義書をアップロード or プリセット選択
2. アウトプット: マッピングファイルをアップロード or テキスト入力
3. Claude APIがデータパレット構築手順書を生成
4. Excel形式でダウンロード
"""
import csv
import io
import json
import yaml
import base64
import os
import re
import uuid
from datetime import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
except ImportError:
    pass  # Render環境では環境変数で設定

import anthropic
from fastapi import FastAPI, HTTPException, Request, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, Response, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel

from .parser import (
    parse_input_excel, parse_input_csv,
    parse_output_excel, parse_output_csv,
)
from .excel_builder import build_spreadsheet
from .template_engine import generate_procedure_text, render_step

from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded


def get_client_id(request: Request) -> str:
    """X-Client-ID ヘッダーがあればそれを、なければIPアドレスをレートリミットキーとする"""
    return request.headers.get("X-Client-ID") or get_remote_address(request)


limiter = Limiter(key_func=get_client_id)

app = FastAPI(title="データパレット構築手順書ジェネレータ")
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[os.environ.get("CORS_ORIGIN", "*")],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def log_client_ip(request: Request, call_next):
    xff = request.headers.get("x-forwarded-for", "")
    real_ip = xff.split(",")[0].strip() if xff else (request.client.host if request.client else "-")
    print(f"[CLIENT_IP] {real_ip} {request.method} {request.url.path}", flush=True)
    return await call_next(request)


# --- 認証 ---
AUTH_TOKEN = os.environ.get("APP_AUTH_TOKEN", "")
security = HTTPBearer(auto_error=False)


async def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Bearer Token認証。APP_AUTH_TOKEN未設定の場合は認証スキップ（ローカル開発用）"""
    if not AUTH_TOKEN:
        return  # トークン未設定ならスキップ（ローカル）
    if not credentials or credentials.credentials != AUTH_TOKEN:
        raise HTTPException(status_code=401, detail="認証エラー: 無効なトークンです")


# --- パス定義 ---
BASE_DIR = Path(__file__).parent.parent
UPLOAD_DIR = Path(__file__).parent / "uploads"
OUTPUT_DIR = Path(__file__).parent / "output"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# --- RDS PostgreSQL接続 ---
import psycopg2
from psycopg2.extras import RealDictCursor, Json
from psycopg2.pool import ThreadedConnectionPool
from contextlib import contextmanager

DATABASE_URL = os.environ.get("DATABASE_URL", "")
_db_pool = None


def _get_pool():
    global _db_pool
    if _db_pool is None and DATABASE_URL:
        _db_pool = ThreadedConnectionPool(1, 10, DATABASE_URL)
    return _db_pool


@contextmanager
def _db_conn():
    pool = _get_pool()
    if pool is None:
        yield None
        return
    conn = pool.getconn()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


_DDL_STATEMENTS = [
    """
    CREATE TABLE IF NOT EXISTS industries (
        id TEXT PRIMARY KEY,
        label TEXT,
        description TEXT,
        tables JSONB,
        created_at TIMESTAMPTZ DEFAULT NOW(),
        updated_at TIMESTAMPTZ DEFAULT NOW()
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS strategy_templates (
        id TEXT PRIMARY KEY,
        name TEXT,
        description TEXT,
        who TEXT,
        what TEXT,
        "when" TEXT,
        exclude JSONB,
        ask_user JSONB,
        processing_notes JSONB,
        output_columns JSONB,
        output_columns_note TEXT,
        input_tables JSONB,
        created_at TIMESTAMPTZ DEFAULT NOW(),
        updated_at TIMESTAMPTZ DEFAULT NOW()
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS knowledge (
        id TEXT PRIMARY KEY,
        type TEXT,
        strategy_name TEXT,
        strategy_summary TEXT,
        output_columns JSONB,
        input_table_names JSONB,
        correction TEXT,
        summary TEXT,
        processing_steps JSONB,
        processing_groups JSONB,
        created_at TIMESTAMPTZ DEFAULT NOW(),
        raw_data JSONB
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS sessions (
        id TEXT PRIMARY KEY,
        mode TEXT,
        data JSONB,
        created_at TIMESTAMPTZ DEFAULT NOW(),
        updated_at TIMESTAMPTZ DEFAULT NOW()
    )
    """,
    "CREATE INDEX IF NOT EXISTS idx_sessions_created_at ON sessions(created_at)",
    """
    CREATE TABLE IF NOT EXISTS feedback_log (
        id BIGSERIAL PRIMARY KEY,
        session_id TEXT,
        phase TEXT,
        aspect TEXT,
        comment TEXT,
        output_text TEXT,
        card_context JSONB,
        created_at TIMESTAMPTZ DEFAULT NOW()
    )
    """,
    "ALTER TABLE feedback_log ADD COLUMN IF NOT EXISTS aspect TEXT",
    "ALTER TABLE feedback_log ADD COLUMN IF NOT EXISTS output_text TEXT",
    "CREATE INDEX IF NOT EXISTS idx_feedback_log_created_at ON feedback_log(created_at)",
    "CREATE INDEX IF NOT EXISTS idx_feedback_log_session_id ON feedback_log(session_id)",
    "ALTER TABLE industries ADD COLUMN IF NOT EXISTS deleted_at TIMESTAMPTZ",
]


def _init_db_schema():
    with _db_conn() as conn:
        if conn is None:
            return
        with conn.cursor() as cur:
            for ddl in _DDL_STATEMENTS:
                cur.execute(ddl)


@app.on_event("startup")
async def _on_startup():
    try:
        _init_db_schema()
    except Exception as e:
        print(f"DB schema init failed: {e}")


# (python_key, sql_column) — "when" は予約語のため二重引用符
_INDUSTRY_COLS = [
    ("id", "id"),
    ("label", "label"),
    ("description", "description"),
    ("tables", "tables"),
    ("created_at", "created_at"),
    ("updated_at", "updated_at"),
]

_STRATEGY_COLS = [
    ("id", "id"),
    ("name", "name"),
    ("description", "description"),
    ("who", "who"),
    ("what", "what"),
    ("when", '"when"'),
    ("exclude", "exclude"),
    ("ask_user", "ask_user"),
    ("processing_notes", "processing_notes"),
    ("output_columns", "output_columns"),
    ("output_columns_note", "output_columns_note"),
    ("input_tables", "input_tables"),
    ("created_at", "created_at"),
    ("updated_at", "updated_at"),
]


def _upsert(table: str, entry: dict, columns: list[tuple[str, str]]) -> None:
    present = [(pk, sc) for pk, sc in columns if pk in entry]
    if not present:
        return
    sql_cols = [sc for _, sc in present]
    placeholders = ["%s"] * len(present)
    updates = [f"{sc} = EXCLUDED.{sc}" for _, sc in present if sc != "id"]
    update_clause = ", ".join(updates) if updates else "id = EXCLUDED.id"
    sql = (
        f"INSERT INTO {table} ({', '.join(sql_cols)}) "
        f"VALUES ({', '.join(placeholders)}) "
        f"ON CONFLICT (id) DO UPDATE SET {update_clause}"
    )
    values = []
    for pk, _ in present:
        v = entry[pk]
        values.append(Json(v) if isinstance(v, (dict, list)) else v)
    with _db_conn() as conn:
        if conn is None:
            raise RuntimeError("DB未設定")
        with conn.cursor() as cur:
            cur.execute(sql, values)


def _update_row(table: str, row_id: str, update: dict, columns: list[tuple[str, str]]) -> None:
    col_map = {pk: sc for pk, sc in columns}
    present = [(k, col_map[k]) for k in update if k in col_map]
    if not present:
        return
    set_clause = ", ".join(f"{sc} = %s" for _, sc in present)
    values = []
    for pk, _ in present:
        v = update[pk]
        values.append(Json(v) if isinstance(v, (dict, list)) else v)
    values.append(row_id)
    with _db_conn() as conn:
        if conn is None:
            raise RuntimeError("DB未設定")
        with conn.cursor() as cur:
            cur.execute(f"UPDATE {table} SET {set_clause} WHERE id = %s", values)


def _delete_row(table: str, row_id: str) -> None:
    with _db_conn() as conn:
        if conn is None:
            raise RuntimeError("DB未設定")
        with conn.cursor() as cur:
            cur.execute(f"DELETE FROM {table} WHERE id = %s", (row_id,))


# --- 業界プリセット ---
INDUSTRIES_PATH = BASE_DIR / "templates" / "industries.json"
with open(INDUSTRIES_PATH, encoding="utf-8") as f:
    _INDUSTRIES_FILE = json.load(f).get("industries", {})


def _load_industries() -> dict:
    """業界プリセットをRDSから読み込み、フォールバックでJSONファイル"""
    try:
        with _db_conn() as conn:
            if conn is not None:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("SELECT * FROM industries WHERE deleted_at IS NULL")
                    rows = cur.fetchall()
                    if rows:
                        return {row["id"]: dict(row) for row in rows}
    except Exception:
        pass
    return _INDUSTRIES_FILE


# 起動時の初期値（API呼び出し時に毎回DBから読む）
INDUSTRIES = _INDUSTRIES_FILE

# --- セッション管理（RDS バックエンド） ---
import threading
import time as _time


class SessionStore:
    """RDS-backed session store with dict-like API.

    Use `with sessions.transaction(sid) as session:` for read+mutate flows
    so the dict is auto-saved on context exit.
    """

    def __contains__(self, sid: str) -> bool:
        with _db_conn() as conn:
            if conn is None:
                return False
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM sessions WHERE id = %s", (sid,))
                return cur.fetchone() is not None

    def __getitem__(self, sid: str) -> dict:
        data = self.get(sid)
        if data is None:
            raise KeyError(sid)
        return data

    def __setitem__(self, sid: str, data: dict) -> None:
        self.save(sid, data)

    def get(self, sid: str, default=None):
        with _db_conn() as conn:
            if conn is None:
                return default
            with conn.cursor() as cur:
                cur.execute("SELECT data FROM sessions WHERE id = %s", (sid,))
                row = cur.fetchone()
                if row is None:
                    return default
                return row[0]

    def save(self, sid: str, data: dict) -> None:
        mode = data.get("mode") if isinstance(data, dict) else None
        with _db_conn() as conn:
            if conn is None:
                raise RuntimeError("DB未設定")
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO sessions (id, mode, data, updated_at)
                    VALUES (%s, %s, %s, NOW())
                    ON CONFLICT (id) DO UPDATE SET
                        mode = EXCLUDED.mode,
                        data = EXCLUDED.data,
                        updated_at = NOW()
                    """,
                    (sid, mode, Json(data)),
                )

    def pop(self, sid: str, default=None):
        with _db_conn() as conn:
            if conn is None:
                return default
            with conn.cursor() as cur:
                cur.execute("DELETE FROM sessions WHERE id = %s RETURNING data", (sid,))
                row = cur.fetchone()
                return row[0] if row else default

    @contextmanager
    def transaction(self, sid: str):
        """Load session, yield it, save on exit. Yields None if not found."""
        data = self.get(sid)
        if data is None:
            yield None
            return
        yield data
        self.save(sid, data)

    def cleanup_older_than(self, hours: int) -> int:
        with _db_conn() as conn:
            if conn is None:
                return 0
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM sessions WHERE created_at < NOW() - make_interval(hours => %s)",
                    (hours,),
                )
                return cur.rowcount


sessions = SessionStore()

# --- ファイル自動クリーンアップ ---


def _cleanup_old_files():
    """1時間ごとにアップロード/出力ファイルをクリーンアップ（セッションはRDSに90日保持）"""
    FILE_MAX_AGE_HOURS = 24
    SESSION_MAX_AGE_DAYS = 90
    while True:
        _time.sleep(3600)  # 1時間ごとに実行
        now = _time.time()
        for d in [UPLOAD_DIR, OUTPUT_DIR]:
            for f in d.iterdir():
                if f.is_file() and (now - f.stat().st_mtime) > FILE_MAX_AGE_HOURS * 3600:
                    try:
                        f.unlink()
                    except OSError:
                        pass
        try:
            sessions.cleanup_older_than(SESSION_MAX_AGE_DAYS * 24)
        except Exception as e:
            print(f"session cleanup failed: {e}")


_cleanup_thread = threading.Thread(target=_cleanup_old_files, daemon=True)
_cleanup_thread.start()

# --- Claude APIクライアント ---
client = anthropic.Anthropic(max_retries=1, timeout=60.0)
async_client = anthropic.AsyncAnthropic(max_retries=1, timeout=180.0)

# --- データモデル ---
class GenerateRequest(BaseModel):
    session_id: str | None = None
    input_tables: list[dict]  # [{table_name, columns: [{name, type, description}]}]
    output_mapping: dict      # {columns: [{name, definition, source_column, source_table}]}
    additional_context: str = ""

class ChatRequest(BaseModel):
    session_id: str
    message: str

class ChatResponse(BaseModel):
    session_id: str
    reply: str
    status: str  # "asking" | "done" | "review" | "consultation_complete"
    download_url: str | None = None
    consultation_result: dict | None = None


class ConsultationStartRequest(BaseModel):
    message: str
    industry: str | None = None


class ConsultationApplyRequest(BaseModel):
    session_id: str
    input_tables: list[dict] | None = None
    output_mapping: dict | None = None


class OrganizationStartRequest(BaseModel):
    consultation_session_id: str
    input_tables: list[dict]
    additional_hint: str | None = None


class OrganizationChatRequest(BaseModel):
    session_id: str
    message: str


class OrganizationUpdateTablesRequest(BaseModel):
    session_id: str
    input_tables: list[dict]


class OrganizationFinalizeRequest(BaseModel):
    session_id: str


class FeedbackRequest(BaseModel):
    session_id: str
    result: str  # "ok" | "fix"
    fix_description: str = ""


# --- ナレッジベース管理（Supabase or JSONファイル） ---
KNOWLEDGE_PATH = BASE_DIR / "skills" / "knowledge_base.json"


def _load_knowledge_base() -> list[dict]:
    """ナレッジ全件取得"""
    try:
        with _db_conn() as conn:
            if conn is not None:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute(
                        "SELECT * FROM knowledge ORDER BY created_at DESC LIMIT 200"
                    )
                    return [dict(row) for row in cur.fetchall()]
    except Exception:
        pass
    if KNOWLEDGE_PATH.exists():
        with open(KNOWLEDGE_PATH, encoding="utf-8") as f:
            return json.load(f)
    return []


def _save_knowledge(entry: dict):
    """ナレッジ1件保存"""
    entry["created_at"] = datetime.now().isoformat()
    entry["id"] = str(uuid.uuid4())[:8]

    # 配列/dict以外のフィールドを raw_data にまとめて保存
    db_entry = {
        "id": entry["id"],
        "type": entry.get("type", ""),
        "strategy_name": entry.get("strategy_name", ""),
        "strategy_summary": entry.get("strategy_summary", ""),
        "output_columns": entry.get("output_columns", []),
        "input_table_names": entry.get("input_table_names", []),
        "correction": entry.get("correction", ""),
        "summary": entry.get("summary", ""),
        "processing_steps": entry.get("processing_steps", []),
        "processing_groups": entry.get("processing_groups", []),
        "created_at": entry["created_at"],
        "raw_data": entry,  # 全フィールドをJSONBに保存（後方互換）
    }

    try:
        with _db_conn() as conn:
            if conn is not None:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO knowledge
                        (id, type, strategy_name, strategy_summary, output_columns,
                         input_table_names, correction, summary, processing_steps,
                         processing_groups, created_at, raw_data)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            db_entry["id"], db_entry["type"], db_entry["strategy_name"],
                            db_entry["strategy_summary"], Json(db_entry["output_columns"]),
                            Json(db_entry["input_table_names"]), db_entry["correction"],
                            db_entry["summary"], Json(db_entry["processing_steps"]),
                            Json(db_entry["processing_groups"]), db_entry["created_at"],
                            Json(db_entry["raw_data"]),
                        ),
                    )
                    return
    except Exception:
        pass

    kb = []
    if KNOWLEDGE_PATH.exists():
        with open(KNOWLEDGE_PATH, encoding="utf-8") as f:
            kb = json.load(f)
    kb.append(entry)
    if len(kb) > 200:
        kb = kb[-200:]
    with open(KNOWLEDGE_PATH, "w", encoding="utf-8") as f:
        json.dump(kb, f, ensure_ascii=False, indent=2)


def _delete_knowledge(entry_id: str) -> bool:
    """ナレッジ1件削除"""
    try:
        with _db_conn() as conn:
            if conn is not None:
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM knowledge WHERE id = %s", (entry_id,))
                    return cur.rowcount > 0
    except Exception:
        pass
    if KNOWLEDGE_PATH.exists():
        with open(KNOWLEDGE_PATH, encoding="utf-8") as f:
            kb = json.load(f)
        new_kb = [e for e in kb if e.get("id") != entry_id]
        if len(new_kb) < len(kb):
            with open(KNOWLEDGE_PATH, "w", encoding="utf-8") as f:
                json.dump(new_kb, f, ensure_ascii=False, indent=2)
            return True
    return False


def _update_knowledge(entry_id: str, update: dict) -> dict | None:
    """ナレッジ1件更新"""
    update.pop("id", None)
    update.pop("created_at", None)

    try:
        with _db_conn() as conn:
            if conn is not None:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    if "raw_data" not in update:
                        cur.execute(
                            "SELECT raw_data FROM knowledge WHERE id = %s",
                            (entry_id,),
                        )
                        row = cur.fetchone()
                        if row:
                            raw = row["raw_data"] or {}
                            if isinstance(raw, str):
                                raw = json.loads(raw)
                            raw.update(update)
                            update["raw_data"] = raw
                    set_parts = []
                    values = []
                    for k, v in update.items():
                        set_parts.append(f"{k} = %s")
                        values.append(Json(v) if isinstance(v, (dict, list)) else v)
                    values.append(entry_id)
                    cur.execute(
                        f"UPDATE knowledge SET {', '.join(set_parts)} "
                        f"WHERE id = %s RETURNING *",
                        values,
                    )
                    row = cur.fetchone()
                    return dict(row) if row else None
    except Exception:
        pass
    if KNOWLEDGE_PATH.exists():
        with open(KNOWLEDGE_PATH, encoding="utf-8") as f:
            kb = json.load(f)
        for i, e in enumerate(kb):
            if e.get("id") == entry_id:
                kb[i].update(update)
                with open(KNOWLEDGE_PATH, "w", encoding="utf-8") as f:
                    json.dump(kb, f, ensure_ascii=False, indent=2)
                return kb[i]
    return None


def get_similar_knowledge(output_mapping: dict, limit: int = 3) -> str:
    """アウトプット定義から類似ナレッジを検索"""
    kb = _load_knowledge_base()
    if not kb:
        return ""
    output_keywords = set()
    for col in output_mapping.get("columns", []):
        for val in [col.get("name", ""), col.get("definition", "")]:
            for word in val.split():
                if len(word) > 1:
                    output_keywords.add(word)
    scored = []
    for entry in kb:
        score = 0
        entry_text = json.dumps(entry, ensure_ascii=False, default=str)
        for kw in output_keywords:
            if kw in entry_text:
                score += 1
        if score > 0:
            scored.append((score, entry))
    scored.sort(key=lambda x: -x[0])
    if not scored:
        return ""
    lines = ["## 類似ケースのナレッジ（過去の成功事例）"]
    for _, entry in scored[:limit]:
        lines.append(f"\n### {entry.get('summary', '不明')}")
        steps = entry.get("processing_steps", [])
        for s in steps[:5]:
            op = s.get("operation", "")
            lines.append(f"- {op}: {json.dumps(s.get('settings', {}), ensure_ascii=False)[:100]}")
        if entry.get("fix_description"):
            lines.append(f"- ※修正あり: {entry['fix_description']}")
    return "\n".join(lines)


# --- スキルローダー ---
SKILLS_DIR = BASE_DIR / "skills"


def load_skill(name: str) -> str:
    """スキルファイルを読み込む（.md優先、なければ.yaml、patterns/配下も対応）"""
    # まず直下を探す
    for ext in (".md", ".yaml"):
        path = SKILLS_DIR / f"{name}{ext}"
        if path.exists():
            return path.read_text(encoding="utf-8")
    # patterns/配下を探す
    for ext in (".md", ".yaml"):
        path = SKILLS_DIR / "patterns" / f"{name}{ext}"
        if path.exists():
            return path.read_text(encoding="utf-8")
    return ""


def select_skills_phase1(input_tables: list[dict], output_mapping: dict) -> list[str]:
    """Phase1用: 必要なスキルだけ選択（procedure_formatは除外）"""
    skills = ["base_operations", "design_patterns", "hearing_defaults"]

    # アウトプットのテキストを結合
    output_text = ""
    for col in output_mapping.get("columns", []):
        output_text += col.get("name", "") + " " + col.get("definition", "") + " "

    # インプットのテーブル名
    input_text = " ".join(t.get("table_name", "") for t in input_tables)
    all_text = output_text + input_text

    # web行動キーワード
    web_keywords = ["閲覧", "カート", "かご", "お気に入り", "PV", "Click", "サイト"]
    if any(kw in output_text for kw in web_keywords) or "webアクセスログ" in input_text:
        skills.append("web_data")
        skills.append("web_access")

    # カート・お気に入り → 購入除外
    if any(kw in all_text for kw in ["カート", "かご", "お気に入り"]):
        skills.append("purchase_exclusion")
        skills.append("cart_and_purchase")

    # 誕生日パターン
    if any(kw in all_text for kw in ["誕生日", "生年月日"]):
        skills.append("birthday_pattern")

    return list(dict.fromkeys(skills))


# --- 施策相談エージェント用プロンプト ---
SYSTEM_PROMPT_CONSULTATION = """あなたはb→dashを使ったマーケティング施策の設計コンサルタントです。
ユーザーが実現したい施策を聞き取り、b→dashデータパレットで必要な「インプットテーブル」と「アウトプット定義」を設計します。

## 利用可能なデータテーブル
{industry_tables}

## 過去の施策相談ナレッジ
{knowledge}

## 暗黙の前提（毎回適用・質問しない）
以下は**質問せず常にこの前提**で設計する。明示的にユーザーから別の指定があった場合のみ上書きする。
- **集計単位**: 顧客ID単位で1レコード（1行=1顧客）。質問しない
- **タイムゾーン**: JST（日本標準時）固定
- **週次の起点**: 月曜起点（週次配信・週次集計は月曜〜日曜で切る）
- **「閲覧」の定義**: 商品詳細ページの閲覧を指す。それ以外の閲覧（リスト画面、検索結果画面等）を対象にしたい場合のみユーザーに確認する
- **配信チャネル別の宛先カラム（差し込み項目に自動で入れる）**:
  - メール → メールアドレス
  - LINE → LINE ID
  - SMS → 電話番号
  - Push通知 → FCMトークン
- **配信履歴による除外**: b→dashの「メール行動ログデータ」テーブルを使って判定する（過去N日以内に配信済みの除外はこのテーブルで実現）

## 施策設計の4観点（必ず全て聞ききる）
以下の4つを**全て具体的に確定**させてからアウトプット定義JSONを出力する。1つでも未確定ならJSON出力禁止。

① **誰に送るか**（セグメント条件）: どんな顧客を対象にするか
② **何を送るか / 何を差し込むか**（パーソナライズ情報）: メールやメッセージに何を載せるか
③ **いつ送るか**（配信タイミング）: トリガー/バッチ、頻度、起点
④ **誰には送らないか**（除外条件）: どんな顧客を対象外にするか

**対話の進め方:**
1. ユーザーの最初のメッセージから読み取れる観点は改めて聞かない
2. 読み取れなかった観点を①→②→③→④の順に1つずつ聞いていく
3. 4つ全て確定したら、セルフチェックリストを確認
4. セルフチェックもOKなら、初めてJSON出力する

### ②の聞き方（デフォルトセットに過不足あるかを聞く）
デフォルトの差し込み項目:
- 顧客: 顧客名、メールアドレス、メルマガ配信許可フラグ
- 商品: 商品名、商品画像URL、商品価格、商品詳細URL
- 選択肢: A)デフォルトOK  B)追加あり  C)一部不要

## 基本ルール
- **1ターンに質問は1つだけ**
- 数値（日数・件数・期間）は推論で決めず必ずユーザー確認
- 全ての質問は選択肢A/B/C+「その他」形式。オープンクエスチョン禁止
- 業界に関する質問は絶対にしない

## ビジネス用語→データ条件の変換確認（最重要ルール）
definitionにビジネス用語で要件を書いたら、**それが実際にどのデータのどの値で判定できるのか**を必ずユーザーに確認する。
言葉で定義できても「じゃあそれはどのテーブルのどのカラムがどんな値のときなのか」が不明なままdefinitionに書かない。

**原則**: definitionの全ての条件が「テーブル名.カラム名 = 値」または「テーブル名.カラム名 に '文字列' を含む」のレベルまで落とし込まれていること。

**よくある変換パターン（これらが出たら必ずデータ条件を確認する）:**
| ビジネス用語 | 確認すべきデータ条件 |
|---|---|
| 「商品を閲覧した」 | → webアクセスログで**何のURLパターン**なら商品詳細ページか |
| 「カートに入れた」 | → **何のURLパターン or イベント**でカート投入と判定するか |
| 「購入した」 | → **どのテーブルのどのステータス**が購入完了か。キャンセル含むか |
| 「配信停止している」 | → **どのカラムがどの値**なら停止か |
| 「新規会員」 | → **登録日がN日以内？初回購入前？** 具体的な条件 |
| 「お気に入りに入れた」 | → **何のURLパターン or イベント**で判定するか |
| 「休眠顧客」 | → **最終購入からN日以上？** 具体的な閾値 |
| 「優良顧客」 | → **購入金額N円以上？購入回数N回以上？** 具体基準 |
| 「リピーター」 | → **購入N回以上？** 具体的な閾値 |

上記以外でも、**ビジネス用語がdefinitionに残っていたら、それをデータ条件に変換する質問を追加する**。

**期間表現の具体化も同様:**
- 「昨日」→「処理日-1日の00:00〜23:59」
- 「直近N日」→「処理日-N日〜処理日-1日」（含む/含まないまで明示）

**設計選択**（Phase1で確定させる、Phase2/3に持ち越さない）:
- 購入判定を受注×受注明細（商品ID単位で正確）にするか webコンバージョン（当日リアルタイムだが顧客単位）にするか
  - 当日配信×購入除外の組合せは**両方のトレードオフを説明した上でユーザーに選ばせる**

## 出力前セルフチェック（該当があれば追加質問）
- **「最新/最後/最初/上位N」** → どの日時/数値カラムで順序付けるか確認
- **「購入」判定** → 受注完了 / 発送完了 / 支払完了 / キャンセル含むか確認
- **顧客区分**（新規/リピーター/休眠/優良） → 具体的な閾値を確認
- **差し込み商品** → 在庫・公開・価格帯の条件を確認
- **複数商品を差し込む場合** → 選ぶ優先順位を確認（例: 閲覧日時の新しい順？価格の高い順？閲覧回数の多い順？）
- **異なるロジックの商品グループが複数ある場合**（例: 閲覧商品2件+レコメンド商品3件）→ **グループごとに**件数・選択ロジック・優先順位を個別に確認する。まとめて聞かない。

## レコメンド系施策の特別ルール
「レコメンド」「おすすめ商品」「関連商品」等が出たら、**アウトプット定義を設計する前に**以下をユーザー確認:
1. **レコメンドの起点**: 閲覧商品 / 購買履歴（協調フィルタリング） / 人気ランキング / 顧客属性 / 手動設定リスト
2. **フィルタ**: 在庫・価格帯・既購入除外
3. **件数**、**フォールバック**（ロジック不成立時の代替）

## アウトプットテーブルの設計方針
- **顧客単位の横持ち**（1行=1顧客）。複数商品は「商品名_1, 商品名_2, ...」で横展開
- カラムは2種類: `segment`（条件判定用）/ `personalize`（差し込み用）

## 出力JSON形式（全情報が揃ったら ```json ... ``` で囲んで出力）
```json
{{{{
  "action": "consultation_result",
  "strategy_name": "...",
  "strategy_summary": "...",
  "requirements": {{{{ "who": "...", "what": "...", "when": "...", "exclude": "..." }}}},
  "input_tables": [{{{{"table_name": "...", "columns": [{{{{"name": "...", "type": "...", "description": "..."}}}}]}}}}],
  "output_mapping": {{{{
    "columns": [
      {{{{"name": "...", "definition": "具体的に", "source_column": "...", "source_table": "...", "purpose": "segment|personalize"}}}}
    ]
  }}}}
}}}}
```
- input_tables / output_mapping の source_table / source_column は必ず埋める
- definitionは具体値で書く（悪: 「カートに入れた商品名」／良: 「webアクセスログから過去N日以内の商品ID→商品テーブルと結合→最新投入順3件を横展開」）

## 施策テンプレート（合致すれば output_columns をベースに使う）
{strategy_templates}
- `ask_user` 項目は必ずユーザーに質問（AIが決めない）
- `_N` はユーザー指定の件数に展開
- 合致しなければ4観点ヒアリングからフルで組み立てる
"""


def _load_strategy_templates() -> list[dict]:
    """施策テンプレートを読み込む（RDS → JSONファイルフォールバック）"""
    try:
        with _db_conn() as conn:
            if conn is not None:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("SELECT * FROM strategy_templates")
                    rows = cur.fetchall()
                    if rows:
                        return [dict(row) for row in rows]
    except Exception:
        pass
    tpl_path = SKILLS_DIR / "strategy_templates.json"
    if tpl_path.exists():
        with open(tpl_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def _format_strategy_templates(templates: list[dict]) -> str:
    """テンプレートをプロンプト用テキストに変換"""
    if not templates:
        return "（テンプレートはありません。4観点ヒアリングからフルで組み立ててください。）"
    lines = ["以下の施策テンプレートが利用可能です:\n"]
    for tpl in templates:
        lines.append(f"### {tpl['name']}（ID: {tpl['id']}）")
        lines.append(f"概要: {tpl['description']}")
        lines.append(f"- 誰に: {tpl['who']}")
        what = tpl['what']
        lines.append(f"- 何を: {', '.join(what) if isinstance(what, list) else what}")
        lines.append(f"- いつ: {tpl['when']}")
        lines.append(f"- 除外: {', '.join(tpl['exclude'])}")
        lines.append(f"- **必ずユーザーに聞く項目**: {', '.join(tpl['ask_user'])}")
        notes = tpl.get('processing_notes', [])
        if notes:
            lines.append(f"- **処理上の注意点**:")
            for note in notes:
                lines.append(f"  - {note}")
        lines.append("")
    return "\n".join(lines)


def get_consultation_system_prompt(industry: str | None = None) -> str:
    """施策相談エージェント用のシステムプロンプトを生成"""
    industries = _load_industries()
    if industry and industry in industries:
        ind = industries[industry]
        tables_text = f"**業界: {ind['label']}（{ind['description']}）**\n利用可能なテーブル:\n"
        # tables は dict or list のどちらか
        tbls = ind.get("tables") or ind.get("data_tables") or []
        if isinstance(tbls, dict):
            tbl_items = list(tbls.items())
        else:
            tbl_items = [(t.get("table_name", ""), t) for t in tbls]
        for tbl_name, tbl_info in tbl_items:
            desc = tbl_info.get("description", "") if isinstance(tbl_info, dict) else ""
            col_count = len(tbl_info.get("columns", [])) if isinstance(tbl_info, dict) else 0
            tables_text += f"- **{tbl_name}**（{col_count}カラム）: {desc}\n"
    else:
        tables_text = (
            "業界未選択。以下の汎用EC/マーケティングテーブル構造を前提に設計する:\n"
            "- **顧客**: 顧客ID, メール, 氏名, 生年月日, 配信許可フラグ 等\n"
            "- **受注 / 受注明細**: 購買履歴（顧客ID + 商品ID + 日時）\n"
            "- **商品**: 商品ID, 商品名, 価格, 画像URL, 詳細URL 等\n"
            "- **webアクセスログ**: ビジターID, PV/Click日時, ページURL, リレーション項目_1〜N\n"
            "- **webコンバージョン**: ビジターID, CV時刻, CV名（テナント固有値）\n"
        )

    knowledge_text = _get_consultation_knowledge()
    templates = _load_strategy_templates()
    templates_text = _format_strategy_templates(templates)

    return SYSTEM_PROMPT_CONSULTATION.format(
        industry_tables=tables_text,
        knowledge=knowledge_text or "（過去のナレッジはありません）",
        strategy_templates=templates_text,
    )


# --- テーブル定義整理エージェント用プロンプト ---
SYSTEM_PROMPT_ORGANIZATION = """あなたはb→dashの「テーブル定義整理エージェント」です。
施策相談フェーズで決まったアウトプット定義の各カラムが、**ユーザーが提供した実テーブルのどのカラムから取れるか**を紐付ける、それだけが役割です。

## 施策要件サマリー（Phase1の結果・参照用）
施策名: {strategy_name}
概要: {strategy_summary}

要件:
- 誰に: {who}
- 何を: {what}
- いつ: {when}
- 除外: {exclude}

## Phase1のアウトプット定義（これをマッピング対象にする）
{abstract_mapping}

## ユーザーが提供した実テーブル定義
{actual_tables}

## あなたの仕事（これだけ）
アウトプット定義の各カラムに対して、「実テーブルに必要なデータが揃っているか」を検証し、足りなければ missing として警告する。
可能なら実テーブルのテーブル名/カラム名を source_table / source_column にマッピングする。

- high: 実テーブルに明確に該当カラムがある → 確定
- medium: カラム名が違うが意味的に該当する → 確定として扱い reason に根拠を書く
- low: 判断に迷う → 質問する（ただし下記「聞いていいこと」のスコープのみ）
- なし: 実テーブルに該当カラムが存在しない → missing として警告

## 絶対に聞いてはいけないこと
- ❌ **テナント固有値の確認**（CV名、URLパターン、ステータス値、フラグ値等）→ **Phase1で確定済み**。Phase1のdefinitionに具体値が書かれているので、それをそのまま使う
- ❌ **設計判断**（購入除外を受注データ/webCVどちらで行うか等）→ Phase1で決済み
- ❌ **リレーション項目の特定**（商品IDがリレーション項目_7か_10か等のwebアクセスログ固有のカラム特定）→ **Phase3の手順書生成フェーズで技術確認する**
- ❌ **処理ロジック**（結合、集約、順序付け、分割、置換、型変換）→ Phase3
- ❌ **件数/期間/閾値**（N日以内、上位N件等）→ Phase1で決済み
- ❌ **配信頻度や運用面の判断**

## 聞いていいこと（これだけ）
実テーブルに**明らかに複数の候補カラムがあり、デフォルトで決まらず、かつ Phase3 の技術確認とは性質が違う**場合のみ。ほとんどの場合は質問不要で、デフォルトで埋められるはず。

具体例:
- 顧客テーブルに「姓」「名」「full_name」が全て存在し、どちらを使うか一意に決まらない場合のみ
- （ただし通常は「姓 + 名」で結合するのがデフォルトなので聞かない）

**原則: Phase2では可能な限り質問せず、デフォルトで埋めて organization_complete を出す。**
**迷うくらいなら質問せずに medium confidence で決め打ちして reason に記録する。**

## デフォルトルール（質問せずにこれで埋める）
- 顧客名 → 顧客テーブルの「姓 + 名」を結合
- メールアドレス → 顧客テーブルの「メールアドレス(メイン)」
- LINE ID → 顧客テーブルの LINE ID カラム
- 電話番号 → 顧客テーブルの電話番号カラム
- メルマガ配信許可フラグ → 顧客テーブルの該当カラム
- 商品名/画像URL/価格/詳細URL → 商品テーブルの該当カラム
- webアクセスログ関連のリレーション項目の特定 → **Phase2ではカラム名まで踏み込まず、「webアクセスログから取得」とだけ書く**（Phase3で技術確認）

## 質問が必要になった場合のルール
- **1度に1つだけ質問する**。複数カラムについて同時に質問カードを並べない
- questions 配列には **最大1要素** まで。2個以上は禁止
- ユーザーの回答を待って、次の質問があれば次のターンで出す

## 質問フォーマット
```
質問の説明文

A) 選択肢1
B) 選択肢2
C) 選択肢3
```

## definition に具体値を埋め込むルール
output_mapping.columns[].definition 列は、Phase3に渡したときに**追加質問なしでそのまま処理できる粒度**で書く:
- ❌ 「webコンバージョンから購入を判定」
- ✅ 「webコンバージョンテーブルで conversion_name = '購入' かつ conversion_time が処理日-1日の0:00〜23:59の顧客を1、それ以外0」
- ❌ 「昨日のログから取得」
- ✅ 「web_log_table で access_time が処理日-1日の0:00〜23:59のレコード」

期間、値、カラム名は全て具体値で書く。曖昧な語（昨日/最近/購入関連）は残さない。

## 質問フォーマット（必ずこの形式）
質問が必要な場合、必ず以下の形式で選択肢を提示する:
```
質問の説明文（どのカラムかを聞く内容のみ）

A) 選択肢1の具体的な内容
B) 選択肢2の具体的な内容
C) 選択肢3の具体的な内容
```
オープンクエスチョンは禁止。ユーザーは「その他」で自由記述も可能。

## 出力JSON

質問が残っている場合（自然文の説明 + 下記JSONを ```json ... ``` で囲む）:
```json
{{{{
  "action": "organization_review",
  "confirmed": [{{{{"name": "...", "source_table": "...", "source_column": "...", "confidence": "high|medium", "reason": "..."}}}}],
  "questions": [{{{{"name": "...", "prompt": "...", "options": ["A) ...", "B) ..."]}}}}],
  "missing": [{{{{"name": "...", "reason": "..."}}}}]
}}}}
```

完了時:
```json
{{{{
  "action": "organization_complete",
  "input_tables": [...],
  "output_mapping": {{{{"columns": [{{{{"name": "...", "definition": "...", "source_table": "...", "source_column": "...", "purpose": "segment|personalize"}}}}]}}}}
}}}}
```

## ルール
- questions は最大1件。questionsが空（全確定）なら即 organization_complete
- definitionはPhase1で書かれた内容をベースに実テーブル情報で具体化（加工ロジックは追加しない）
- input_tablesはユーザー提供分そのまま
"""


def get_organization_system_prompt(session: dict) -> str:
    """テーブル定義整理エージェント用プロンプト生成"""
    cr = session.get("consultation_result") or {}
    actual = session.get("input_tables") or []
    reqs = cr.get("requirements") or {}
    return SYSTEM_PROMPT_ORGANIZATION.format(
        strategy_name=cr.get("strategy_name", "不明"),
        strategy_summary=cr.get("strategy_summary", ""),
        who=reqs.get("who", ""),
        what=reqs.get("what", ""),
        when=reqs.get("when", ""),
        exclude=reqs.get("exclude", ""),
        abstract_mapping=json.dumps(cr.get("output_mapping", {}), ensure_ascii=False, indent=2),
        actual_tables=json.dumps(actual, ensure_ascii=False, indent=2),
    )


def _get_consultation_knowledge(limit: int = 3) -> str:
    """施策相談ナレッジを検索"""
    kb = _load_knowledge_base()
    consultation_entries = [e for e in kb if e.get("type") == "consultation"]
    if not consultation_entries:
        return ""
    recent = consultation_entries[-limit:]
    lines = ["以下は過去の施策相談の結果です:"]
    for entry in recent:
        lines.append(f"\n#### {entry.get('strategy_name', '不明')}")
        lines.append(f"概要: {entry.get('strategy_summary', '')}")
        out_cols = entry.get("output_columns", [])
        if out_cols:
            lines.append(f"アウトプットカラム: {', '.join(out_cols[:5])}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Knowledge-base helpers: operation_cards.yaml / operation_schemas.json
# ---------------------------------------------------------------------------
_CARDS_PATH   = Path(__file__).parent.parent / "skills" / "operation_cards.yaml"
_SCHEMAS_PATH = Path(__file__).parent.parent / "skills" / "operation_schemas.json"


def _build_operation_cards_for_step1() -> str:
    """operation_cards.yaml の全30操作を Step1 用に圧縮したテキストを返す。
    1カード4〜5行。カテゴリ別にセクション分け。起動時1回ロード。"""
    try:
        with open(_CARDS_PATH, encoding="utf-8") as f:
            data = yaml.safe_load(f)
    except Exception as e:
        return f"（操作カードの読み込みに失敗: {e}）"

    sections: dict[str, list[str]] = {"統合": [], "加工": [], "テンプレート": []}
    for op_name, card in data.get("operations", {}).items():
        cat = card.get("カテゴリ", "加工")
        desc = card.get("一言説明", "")
        scenes = card.get("使う場面", [])[:2]
        misuses = card.get("使わない場面_よくある誤用", [])
        diffs = card.get("類似操作との違い", {})

        line_scenes = " / ".join(s.split("→")[0].rstrip("、").strip() for s in scenes)
        line_misuse = misuses[0] if misuses else ""
        bdash_notes = card.get("注意_b→dash固有", [])
        line_bdash = bdash_notes[0] if bdash_notes else ""
        diff_items = list(diffs.items())[:1]
        line_diff = "vs {}: {}".format(
            diff_items[0][0], diff_items[0][1].split("。")[0]
        ) if diff_items else ""

        block = f"【{op_name}】{desc}\n  使う場面: {line_scenes}"
        if line_misuse:
            block += f"\n  注意: {line_misuse}"
        if line_bdash:
            block += f"\n  b→dash: {line_bdash}"
        if line_diff:
            block += f"\n  {line_diff}"
        sections.get(cat, sections["加工"]).append(block)

    parts = []
    for cat, blocks in sections.items():
        if blocks:
            parts.append(f"### {cat}\n" + "\n\n".join(blocks))
    return "\n\n".join(parts)


def _build_settings_schemas_for_step2() -> str:
    """Step2 の settings 仕様テキストを返す。
    operation_schemas.json (10操作・詳細) +
    operation_cards.yaml の残り全操作 (必須settings + 最小実行例) をカバー。
    起動時1回ロード。"""
    try:
        with open(_SCHEMAS_PATH, encoding="utf-8") as f:
            schemas = json.load(f).get("operations", {})
    except Exception as e:
        schemas = {}
        print(f"[WARN] operation_schemas.json load failed: {e}", flush=True)

    try:
        with open(_CARDS_PATH, encoding="utf-8") as f:
            cards = yaml.safe_load(f).get("operations", {})
    except Exception as e:
        cards = {}
        print(f"[WARN] operation_cards.yaml load failed: {e}", flush=True)

    lines: list[str] = []

    # ① operation_schemas.json 収録の10操作（詳細スキーマ）
    for op, schema in schemas.items():
        req_fields = schema.get("required", [])
        props = schema.get("properties", {})
        req_strs = []
        for field in req_fields:
            p = props.get(field, {})
            enums = p.get("enum")
            p_desc = p.get("description", "").split("。")[0][:35]
            if enums:
                req_strs.append(f"{field}（{'/'.join(enums)}）")
            elif p.get("type") == "array":
                items_desc = p.get("items", {}).get("description", "")
                if items_desc:
                    req_strs.append(f"{field}（array: {items_desc[:30]}）")
                else:
                    req_strs.append(f"{field}（array）")
            else:
                req_strs.append(f"{field}（{p_desc}）" if p_desc else field)
        opt_fields = [k for k in props if k not in req_fields]
        constraints = schema.get("x-constraints", [])[:2]
        example = schema.get("x-example", {})

        lines.append(f"【{op}】")
        lines.append(f"  必須: {', '.join(req_strs)}")
        if opt_fields:
            lines.append(f"  任意: {', '.join(opt_fields)}")
        for c in constraints:
            lines.append(f"  制約: {c}")
        lines.append(f"  例: {json.dumps(example, ensure_ascii=False)}")
        lines.append("")

    # ② operation_cards.yaml のうち schemas.json に未収録の操作
    schema_ops = set(schemas.keys())
    for op, card in cards.items():
        if op in schema_ops:
            continue
        req = card.get("必須settings", {})
        min_ex = card.get("最小実行例", {})
        ex_out = {k: v for k, v in min_ex.items() if k != "説明" and k in req}
        if not ex_out:
            ex_out = {k: v.split("（")[0].strip() for k, v in list(req.items())[:3]}
        lines.append(f"【{op}】")
        if req:
            lines.append(f"  必須: {', '.join(req.keys())}")
        lines.append(f"  例: {json.dumps(ex_out, ensure_ascii=False)}")
        lines.append("")

    return "\n".join(lines)


# モジュールロード時に1回だけ生成（再起動で反映）
_OP_CARDS_STEP1      = _build_operation_cards_for_step1()
_SETTINGS_SCHEMAS_STEP2 = _build_settings_schemas_for_step2()


# --- Phase1 Step1: 方針決定プロンプト（軽量、Skills無し） ---
SYSTEM_PROMPT_STEP1 = """あなたはb→dashのデータパレット構築を設計するAIです。
SQLの知識をベースに、アウトプットを作るためにどのb→dash操作を使うか方針を決めてください。

## インプットテーブル
{input_tables}

## アウトプット定義
{output_mapping}

## b→dashで使える操作
{operation_cards}

## やること
1. **まずアウトプット定義の「定義」列をチェック**。以下のキーワードがあるのに具体的なロジックが不足している場合は差し戻す：
   - 「同時閲覧」「相関」「併買」→ 算出方法が不明（例: セッションIDベース？受注IDベース？）
   - 「ランキング」「上位N件」→ 何を基準にランキングするか不明
   - 「レコメンド」「おすすめ」→ レコメンドロジックが不明
   - 「フラグ」「判定」→ 判定条件が不明
   差し戻す場合は、**具体的にどう書き直せばいいか提案**すること。例：
   「定義列に以下のように追記してください：『同一セッション内で購入商品と同時に閲覧された商品をセッションIDベースで算出し、閲覧数の多い順に上位8件を横展開』」
2. 定義が十分な場合、アウトプットの各カラムを実現するために必要な操作を洗い出す
3. SQLとして考えた場合の処理フロー（JOIN順序、WHERE条件、GROUP BY等）を整理
4. それをb→dash操作にマッピング

## 技術確認が必要な場合
webアクセスログがインプットに含まれ、かつアウトプットにweb行動由来カラムがある場合のみ質問OK。
**1回に1つだけ質問する。複数の質問をまとめない。**

質問フォーマット（必ずこの形式で）：
```
質問の説明文

A) 選択肢1の具体的な内容
B) 選択肢2の具体的な内容
C) 選択肢3の具体的な内容
```

質問項目（まとめて1回で聞く）：
- 閲覧/カート等の商品IDが格納されているリレーション項目番号
- そのリレーション項目のデータ形式（1レコード1値 or カンマ区切り複数値）
- 顧客IDが格納されているリレーション項目番号

**質問は最大1回。回答を受けたら次は必ずplan JSONを出力すること。同じ質問を繰り返すのは禁止。**
それ以外の質問（対象期間、件数、優先順位等の要件）は禁止。

## 出力形式
技術確認が不要な場合、以下のJSON形式で出力：
```json
{{"action": "plan", "operations": ["横統合", "絞込み", "名寄せ", "テンプレート 縦持ちを横持ちに変換"], "flow": "受注テーブルと商品テーブルを横統合 → 顧客IDで絞込み → 名寄せ → 商品ランキングを横持ち変換", "needs_web_hearing": false}}
```

### flow フィールドのルール（厳守）
- **1〜3行以内、最大200文字**。`operations` の順序を矢印 (→) で繋いだ骨子のみ書く
- STEP-by-STEPの詳細（条件式・結合キー・フィルタ値・カラム名のリスト等）は **絶対に書かない**。詳細はStep2（設計書生成）の役割
- マークダウン見出し (`###`/`##`)・箇条書き・表・コードブロックは禁止
- `要確認事項`・`注釈`・`最終アウトプットカラム` 等の補足セクションも禁止（必要ならStep2の `special_notes` で扱う）
- flowが長くなると max_tokens に達して JSON が途中で切れ、Step2が起動しなくなる。**簡潔に書くことが必須**

技術確認が必要な場合はテキストで質問（1回のみ、回答後は即plan JSON出力）。
**回答を受け取ったら、その回答を反映して必ずplan JSONを出力すること。追加質問は禁止。**
"""

# --- Phase1 Step2: 詳細設計プロンプト（該当Skillsのみ） ---
SYSTEM_PROMPT_STEP2 = """あなたはb→dashのデータパレット構築の「設計書」を生成するAIです。

## インプットテーブル
{input_tables}

## アウトプット定義
{output_mapping}

## 処理方針（Step1で決定済み）
{plan}

## ナレッジ（該当操作のパターンのみ）
{skills}

## ★最重要★ 要件は聞かない・JSONは必ず完結させる
アウトプット定義の「定義」列をそのまま採用。不足なら推論。質問禁止。

### 不足情報の推論ルール
- 必要なカラムは「マッピング定義書からの逆算」で具体的に推論する
  例: アウトプットに「姓名」があり入力に「姓」「名」がある → 横統合の残すカラムに「姓」「名」を含める
- 以下の曖昧表現は**禁止**:
  * 「その他必要な〜」「など」「等」「必要に応じて〜」
  * 「適切な〜」「相応の〜」「〜等のカラム」
- 推論できない情報は special_notes に明記する
  例: "受注テーブルの重複排除ルールが不明のため、最新行を採用"

### JSON出力ルール（厳守）
1. **JSONは必ず閉じること。途中で切れるのは絶対NG。**
2. **ui_pathは省略**（書かない）
3. **settingsは下記「操作別設定仕様」に従い全フィールドを具体的に記載**：
   統合キーは常に "カラムA = カラムB" 平文形式（オブジェクト形式NG）
{settings_schemas}
4. **result, noteは1行以内**
5. **カラム名変更は1ステップにまとめてchanges配列で**
6. **special_notesは最大3個**

## 設計書の出力形式
以下のJSON形式で出力してください（```json で囲む）。

```json
{{
  "action": "design",
  "version": "2.0",
  "summary": "設計書の概要説明",
  "processing_steps": [
    {{
      "step": 1,
      "operation": "b→dash操作名",
      "settings": {{...}},
      "save_as": "ファイル名",
      "result": "結果の説明",
      "note": "更新設定: しない"
    }}
  ],
  "special_notes": ["注意事項"]
}}
```

## 設計書の出力例
業務シナリオ: 受注テーブル（受注ID/顧客ID/受注日/購買金額）+ 顧客テーブル（顧客ID/姓/名）→ 顧客別最終購買日・購買回数・購買金額合計・姓名

```json
{{
  "action": "design",
  "version": "2.0",
  "summary": "受注テーブルを顧客単位で集約後、顧客マスタから姓・名を横統合し連結して姓名カラムを作成",
  "processing_steps": [
    {{
      "step": 1,
      "operation": "集約",
      "settings": {{
        "まとめる単位": "顧客ID",
        "集約定義": [
          {{"対象項目": "受注日",   "集計方法": "最新",  "出力カラム名": "最終購買日"}},
          {{"対象項目": "受注ID",   "集計方法": "COUNT", "出力カラム名": "購買回数"}},
          {{"対象項目": "購買金額", "集計方法": "SUM",   "出力カラム名": "購買金額合計"}}
        ]
      }},
      "save_as": "顧客別集計",
      "result": "顧客IDごとに最終購買日・購買回数・購買金額合計が1行に集約",
      "note": "更新設定: しない"
    }},
    {{
      "step": 2,
      "operation": "横統合",
      "settings": {{
        "左ファイル": "顧客別集計",
        "右ファイル": "顧客テーブル",
        "統合方法": "左",
        "統合キー": "顧客ID",
        "残すカラム": ["姓", "名"]
      }},
      "save_as": "顧客別集計_姓名付き",
      "result": "集計結果に顧客マスタの姓・名を付与",
      "note": "更新設定: しない"
    }},
    {{
      "step": 3,
      "operation": "連結",
      "settings": {{
        "連結対象": ["姓", "名"]
      }},
      "save_as": "顧客別購買サマリ",
      "result": "姓と名を連結した姓名カラムを追加",
      "note": "更新設定: しない"
    }}
  ],
  "special_notes": [
    "受注テーブルに重複行がある場合はSTEP1前に名寄せで最新行に統合すること"
  ]
}}
```
"""

# 修正再生成時にSYSTEM_PROMPT_STEP2の末尾に追記して修正指示を最優先にする
SYSTEM_PROMPT_REGEN_SUFFIX = """

## ★★★ 修正指示（このセクションを最優先すること） ★★★
直前に出力した設計書JSONを以下の修正指示に従って修正し、完全な設計書JSONを再出力してください。
- 修正指示に明記された箇所のみ変更し、それ以外は直前のJSONから引き継ぐ
- 修正指示と上記の処理方針が矛盾する場合は修正指示を優先する
- JSONは必ず完結させること
- 質問禁止。不明点は最良の推測で補完してJSON出力すること
- 必ず ```json ブロックでJSONを出力すること。テキストのみの返答は禁止

修正指示：
{correction}
"""


def format_input_tables(tables: list[dict]) -> str:
    """インプットテーブル定義を整形テキストにする"""
    lines = []
    for table in tables:
        lines.append(f"### {table['table_name']}テーブル")
        lines.append("| カラム名 | 型 | 説明 |")
        lines.append("|---------|-----|------|")
        for col in table.get("columns", []):
            lines.append(f"| {col['name']} | {col.get('type', '')} | {col.get('description', '')} |")
        lines.append("")
    return "\n".join(lines)


def format_output_mapping(mapping: dict) -> str:
    """アウトプットマッピング定義を整形テキストにする"""
    lines = [
        "| カラム名 | 定義 | インプットカラム | インプットデータ |",
        "|---------|------|----------------|----------------|",
    ]
    for col in mapping.get("columns", []):
        src_col = col.get("source_column", "") or "（加工で生成）"
        src_tbl = col.get("source_table", "") or ""
        lines.append(f"| {col['name']} | {col.get('definition', '')} | {src_col} | {src_tbl} |")
    return "\n".join(lines)


def get_system_prompt_step1(input_tables: list[dict], output_mapping: dict) -> str:
    """Phase1 Step1: 方針決定（カラム詳細含む）"""
    return SYSTEM_PROMPT_STEP1.format(
        input_tables=format_input_tables(input_tables),
        output_mapping=format_output_mapping(output_mapping),
        operation_cards=_OP_CARDS_STEP1,
    )


def get_system_prompt_step2(input_tables: list[dict], output_mapping: dict, plan: str) -> str:
    """Phase1 Step2: 詳細設計（該当Skillsのみ）"""
    # planから操作リストを抽出してSkills選択
    skill_names = select_skills_from_plan(plan, input_tables, output_mapping)
    skills_text = "\n\n---\n\n".join(load_skill(name) for name in skill_names if load_skill(name))

    # Past knowledge (success + correction cases)
    knowledge_text = get_similar_knowledge(output_mapping)
    if knowledge_text:
        skills_text += chr(10)*2 + "---" + chr(10)*2 + knowledge_text

    return SYSTEM_PROMPT_STEP2.format(
        input_tables=format_input_tables(input_tables),
        output_mapping=format_output_mapping(output_mapping),
        plan=plan,
        skills=skills_text,
        settings_schemas=_SETTINGS_SCHEMAS_STEP2,
    )


def select_skills_from_plan(plan: str, input_tables: list[dict], output_mapping: dict) -> list[str]:
    """planの操作リストに基づいて必要なSkillsだけ選択"""
    skills = []

    # design_patterns（共通ルール）は常に読む
    skills.append("design_patterns")

    # アウトプットのテキスト
    output_text = ""
    for col in output_mapping.get("columns", []):
        output_text += col.get("name", "") + " " + col.get("definition", "") + " "
    all_text = plan + " " + output_text

    # インプットのテーブル名
    input_text = " ".join(t.get("table_name", "") for t in input_tables)

    # --- パターン別ファイル選択 ---

    # web系 → web_access + hearing_defaults + web_data
    if "webアクセスログ" in input_text or "web" in plan.lower():
        skills.append("hearing_defaults")
        skills.append("web_data")
        skills.append("web_access")  # patterns/web_access.md

    # カート・お気に入り・購入除外 → cart_and_purchase + purchase_exclusion
    if any(kw in all_text for kw in ["カート", "かご", "お気に入り", "購入除外", "購入済み"]):
        skills.append("purchase_exclusion")
        skills.append("cart_and_purchase")  # patterns/cart_and_purchase.md

    # 誕生日
    if any(kw in all_text for kw in ["誕生日", "生年月日", "年齢"]):
        skills.append("birthday_pattern")

    # 金額・KPI・パーセント・フォーマット
    if any(kw in all_text for kw in ["金額", "売上", "KPI", "パーセント", "受注回数", "四則演算", "期間判定"]):
        skills.append("calculation_and_format")  # patterns/

    # 差分・多段統合・時系列・LTV
    if any(kw in all_text for kw in ["差分", "前日", "多段", "LTV", "月次", "ダミー", "時系列"]):
        skills.append("advanced_integration")  # patterns/

    # URL生成・置換・テキスト操作
    if any(kw in all_text for kw in ["URL生成", "URLエンコード", "正規表現", "置換", "NULL", "型変換"]):
        skills.append("text_and_cleanup")  # patterns/

    # 差し替え・タスク編集・SQLジョブ
    if any(kw in all_text for kw in ["差し替え", "タスク編集", "SQLジョブ", "年月マスタ", "地域分割"]):
        skills.append("operational")  # patterns/

    return list(dict.fromkeys(skills))  # 重複除去（順序保持）


# 後方互換用
def get_system_prompt(input_tables: list[dict], output_mapping: dict) -> str:
    return get_system_prompt_step1(input_tables, output_mapping)


# build_spreadsheet は excel_builder.py からインポート済み
# generate_procedure_text は template_engine.py からインポート済み


def _parse_json_with_repair(json_str: str) -> dict:
    """JSONパース。途中で切れてる場合は修復を試みる"""
    # まず普通にパース
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        pass

    # 戦略1: 最後の完全な閉じ括弧まで切り詰める（最も確実）
    # processing_stepsの最後の完全なstepまでを取得
    for i in range(len(json_str) - 1, 0, -1):
        if json_str[i] in ('}', ']'):
            # そこまでで切って、足りない括弧を補完
            truncated = json_str[:i+1]
            # 開き括弧と閉じ括弧の差分を計算
            open_braces = truncated.count('{') - truncated.count('}')
            open_brackets = truncated.count('[') - truncated.count(']')
            suffix = ']' * open_brackets + '}' * open_braces
            try:
                data = json.loads(truncated + suffix)
                print(f"[DEBUG] JSON repaired by truncating at {i} + suffix {repr(suffix)}")
                return data
            except json.JSONDecodeError:
                continue

    # 戦略2: processing_stepsだけ抽出
    if '"processing_steps"' in json_str:
        try:
            # processing_stepsの開始位置を見つける
            steps_start = json_str.index('"processing_steps"')
            # その前までのヘッダー部分を取得
            header = json_str[:steps_start]
            # processing_steps配列内の最後の完全なオブジェクトを見つける
            rest = json_str[steps_start:]
            bracket_start = rest.index('[')
            steps_content = rest[bracket_start:]

            # 最後の "}," または "}" を見つけて切る
            last_complete = -1
            for j in range(len(steps_content) - 1, 0, -1):
                if steps_content[j] == '}':
                    test = steps_content[:j+1]
                    open_b = test.count('[') - test.count(']')
                    close_suffix = ']' * open_b
                    try:
                        json.loads('{"test":' + test + close_suffix + '}')
                        last_complete = j
                        break
                    except:
                        continue

            if last_complete > 0:
                fixed_steps = steps_content[:last_complete+1]
                open_brackets = fixed_steps.count('[') - fixed_steps.count(']')
                fixed_steps += ']' * open_brackets
                full_json = header + '"processing_steps": ' + fixed_steps + '}'
                data = json.loads(full_json)
                print(f"[DEBUG] JSON repaired via steps extraction")
                return data
        except Exception:
            pass

    # 戦略3: 文字列の途中切れを処理（未閉じのダブルクォートを閉じる）
    # 最後のダブルクォートの位置を確認
    cleaned = json_str.rstrip()
    # 未閉じの文字列を閉じてから再試行
    if cleaned.count('"') % 2 != 0:
        cleaned += '"'
    # 括弧バランス修復
    open_braces = cleaned.count('{') - cleaned.count('}')
    open_brackets = cleaned.count('[') - cleaned.count(']')
    if open_braces > 0 or open_brackets > 0:
        suffix = ']' * max(0, open_brackets) + '}' * max(0, open_braces)
        try:
            data = json.loads(cleaned + suffix)
            print(f"[DEBUG] JSON repaired via quote+bracket fix")
            return data
        except json.JSONDecodeError:
            pass

    raise json.JSONDecodeError("Cannot repair truncated JSON", json_str, 0)


# --- APIエンドポイント ---

@app.get("/api/industries", dependencies=[Depends(verify_token)])
async def list_industries():
    """業界プリセット一覧を返す"""
    industries = _load_industries()
    result = {}
    for key, ind in industries.items():
        # DB形式: tables は直接 JSONB 配列
        tables = ind.get("tables") or []
        # ファイル形式（後方互換）: data_tables dict
        if not tables and "data_tables" in ind:
            tables = [{"table_name": k, "columns": v.get("columns", [])}
                      for k, v in ind["data_tables"].items()]
        result[key] = {
            "label": ind.get("label", ""),
            "description": ind.get("description", ""),
            "tables": tables,
        }
    return result


class UploadRequest(BaseModel):
    filename: str
    content: str  # base64 encoded file content
    file_type: str = "input"


@app.post("/api/parse", dependencies=[Depends(verify_token)])
async def upload_file(req: UploadRequest):
    """ファイルをアップロードして解析結果を返す（base64 JSON）"""
    suffix = Path(req.filename).suffix.lower()
    if suffix not in (".xlsx", ".xls", ".csv"):
        raise HTTPException(400, "対応形式: .xlsx, .csv")

    save_path = UPLOAD_DIR / f"{uuid.uuid4()}{suffix}"
    with open(save_path, "wb") as f:
        f.write(base64.b64decode(req.content))

    try:
        if req.file_type == "input":
            if suffix == ".csv":
                result = parse_input_csv(str(save_path), table_name=Path(req.filename).stem)
            else:
                result = parse_input_excel(str(save_path))
            return {"type": "input", "tables": result, "filename": req.filename}
        else:
            if suffix == ".csv":
                result = parse_output_csv(str(save_path))
            else:
                result = parse_output_excel(str(save_path))
            return {"type": "output", "mapping": result, "filename": req.filename}
    except Exception as e:
        raise HTTPException(400, f"ファイルの解析に失敗しました。ファイル形式を確認してください。")


def _build_from_design_doc(session_id: str, session: dict, generation_data: dict, input_tables: list) -> "ChatResponse":
    """design actionのJSONからExcelを生成してChatResponseを返す（_generate_impl_bodyとregenerateで共用）"""
    session["design_doc"] = generation_data
    try:
        procedure_text = generate_procedure_text(generation_data)
        session["procedure_text"] = procedure_text

        steps = generation_data.get("processing_steps", [])

        proc_rows = []
        step_num = 1
        for s in steps:
            if not isinstance(s, dict):
                continue
            step_val = s.get("step", "")
            sn = str(step_num) if step_val else ""
            if step_val:
                step_num += 1
            op = s.get("operation", "")
            settings = s.get("settings", {})
            if not isinstance(settings, dict):
                settings = {}
            save_as = s.get("save_as", "")

            try:
                template_text = render_step(s)
            except Exception:
                template_text = f"『{op}』"

            param_values = []
            for v in list(settings.values())[:5]:
                if isinstance(v, list):
                    parts = [json.dumps(item, ensure_ascii=False) if isinstance(item, dict) else str(item) for item in v]
                    param_values.append(", ".join(parts))
                elif isinstance(v, dict):
                    param_values.append(json.dumps(v, ensure_ascii=False))
                else:
                    param_values.append(str(v))
            while len(param_values) < 5:
                param_values.append("")

            proc_rows.append([sn, op, "", template_text, save_as, "", *param_values, template_text])

        # フロー図用: インプットテーブル名とアウトプット名をexcel_dataに渡す
        input_names = [t.get("table_name", f"テーブル{i+1}") for i, t in enumerate(input_tables)]
        output_name = next((s.get("save_as", "") for s in reversed(steps) if isinstance(s, dict) and s.get("save_as")), "")

        excel_data = {
            "action": "generate",
            "title": generation_data.get("summary", "データパレット構築手順書"),
            "input_tables": input_names,
            "output_name": output_name,
            "sections": [
                {
                    "sheet_name": "結合・加工",
                    "title": "手順書",
                    "columns": ["対象作業No", "アイコン", "", "アイコン利用方法", "作成後項目名", "", "対象1", "対象2", "対象3", "対象4", "対象5", "完成形テキスト"],
                    "rows": proc_rows,
                }
            ],
        }

        filepath, filename = build_spreadsheet(excel_data)
        session["last_file"] = filepath
        session["last_filename"] = filename

        summary_lines = ["📋 **設計書サマリー**", f"**概要**: {generation_data.get('summary', '')}"]
        summary_lines.append(f"**処理ステップ数**: {len([s for s in steps if isinstance(s, dict) and s.get('step')])}")
        for s in steps:
            if isinstance(s, dict) and s.get("step"):
                summary_lines.append(f"  Step{s['step']}: {s.get('operation', '')} → {s.get('save_as', '')}")
        summary_lines += ["", "📝 **手順書プレビュー**"]
        design_summary = "\n".join(summary_lines)

        design_path = OUTPUT_DIR / f"design_{session_id}.json"
        with open(design_path, "w", encoding="utf-8") as f:
            json.dump(generation_data, f, ensure_ascii=False, indent=2)
        session["design_file"] = str(design_path)

        return ChatResponse(
            session_id=session_id,
            reply=f"{design_summary}\n\n{procedure_text[:3000]}",
            status="done",
            download_url=f"/api/download/{session_id}",
        )
    except Exception as e:
        return ChatResponse(
            session_id=session_id,
            reply=f"Phase3（手順書生成）でエラー: {e}\n\n設計書JSON:\n{json.dumps(generation_data, ensure_ascii=False, indent=2)[:2000]}",
            status="asking",
        )


async def _generate_core(req: GenerateRequest) -> ChatResponse:
    """設計書生成のコアロジック（内部呼び出し用）"""
    return await _generate_impl(req)


class RegenerateRequest(BaseModel):
    session_id: str
    correction: str


@app.post("/api/regenerate", response_model=ChatResponse, dependencies=[Depends(verify_token)])
@limiter.limit("10/minute")
async def regenerate(request: Request, req: RegenerateRequest):
    """修正依頼をもとにPhase3の手順書を再生成する（Step2のみ再実行）"""
    session = sessions.get(req.session_id)
    if not session:
        raise HTTPException(404, "セッションが見つかりません")

    plan = session.get("plan")
    if not plan:
        raise HTTPException(400, "設計書のプランが見つかりません。先に手順書を生成してください。")

    input_tables = session.get("input_tables", [])
    output_mapping = session.get("output_mapping", {})

    msgs2 = session.get("messages_step2", [])
    print(f"[REGEN] session={req.session_id[:8]} messages_step2 count={len(msgs2)}", flush=True)
    for i, m in enumerate(msgs2):
        print(f"[REGEN]   [{i}] role={m.get('role')} content_len={len(m.get('content',''))}", flush=True)
    print(f"[REGEN] correction='{req.correction[:100]}'", flush=True)
    print(f"[REGEN] input_tables count={len(input_tables)} plan_exists={bool(plan)}", flush=True)

    if not msgs2:
        session["messages_step2"] = [
            {"role": "user", "content": "処理方針に基づいて設計書JSONを出力してください。"}
        ]

    correction_msg = f"以下の点を修正して設計書を再生成してください：\n\n{req.correction}"
    session["messages_step2"].append({"role": "user", "content": correction_msg})

    try:
        # 修正指示をsystemプロンプトに追記することでStep2の元制約より優先させる
        step2_prompt = get_system_prompt_step2(input_tables, output_mapping, plan)
        step2_prompt += SYSTEM_PROMPT_REGEN_SUFFIX.format(correction=req.correction)
        response2 = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=16000,
            system=step2_prompt,
            messages=session["messages_step2"],
        )
        assistant_text = response2.content[0].text
        if response2.stop_reason == "max_tokens":
            print(f"[WARN] regenerate Step2 truncated (session={req.session_id})", flush=True)
        has_json = "```json" in assistant_text
        print(f"[REGEN] new response len={len(assistant_text)} has_json={has_json}", flush=True)
        if not has_json:
            print(f"[REGEN] no JSON, AI text={assistant_text[:300]!r}", flush=True)
            session["messages_step2"].pop()
            sessions.save(req.session_id, session)
            return ChatResponse(
                session_id=req.session_id,
                reply=f"修正内容をより具体的に入力してください。\n\nAIの返答：{assistant_text[:300]}",
                status="asking",
            )
        try:
            new_steps = json.loads(assistant_text.split("```json")[1].split("```")[0].strip()).get("processing_steps", [])
            print(f"[REGEN] new design steps={len(new_steps)}: {[s.get('operation') for s in new_steps if isinstance(s, dict)]}", flush=True)
        except Exception:
            pass
        session["messages_step2"].append({"role": "assistant", "content": assistant_text})
    except Exception as e:
        session["messages_step2"].pop()
        sessions.save(req.session_id, session)
        raise HTTPException(500, f"再生成エラー: {e}")

    result = ChatResponse(session_id=req.session_id, reply="設計書JSONが取得できませんでした。", status="asking")

    if "```json" in assistant_text:
        try:
            json_str = assistant_text.split("```json")[1].split("```")[0].strip()
            generation_data = _parse_json_with_repair(json_str)
            if generation_data.get("action") == "design":
                result = _build_from_design_doc(req.session_id, session, generation_data, input_tables)
            elif generation_data.get("action") == "generate":
                filepath, filename = build_spreadsheet(generation_data)
                session["last_file"] = filepath
                session["last_filename"] = filename
                display_text = assistant_text.split("```json")[0].strip() or "手順書を再生成しました。"
                result = ChatResponse(
                    session_id=req.session_id,
                    reply=display_text,
                    status="done",
                    download_url=f"/api/download/{req.session_id}",
                )
        except (json.JSONDecodeError, IndexError) as e:
            result = ChatResponse(session_id=req.session_id, reply=f"JSON解析エラー: {e}", status="asking")

    sessions.save(req.session_id, session)
    return result


@app.post("/api/generate", response_model=ChatResponse, dependencies=[Depends(verify_token)])
@limiter.limit("10/minute")
async def generate(request: Request, req: GenerateRequest):
    """設計書を生成する（2段階API）- エンドポイント"""
    return await _generate_impl(req)


async def _generate_impl(req: GenerateRequest):
    """設計書を生成する（2段階API）"""
    session_id = req.session_id or str(uuid.uuid4())

    session = sessions.get(session_id)
    if session is None:
        session = {
            "messages": [],
            "messages_step2": [],
            "input_tables": req.input_tables,
            "output_mapping": req.output_mapping,
            "step": "step1",  # step1 or step2
            "plan": None,
            "created_at": _time.time(),
        }

    try:
        return await _generate_impl_body(req, session_id, session)
    finally:
        sessions.save(session_id, session)


async def _generate_impl_body(req: GenerateRequest, session_id: str, session: dict):
    # デバッグログ
    print(f"[DEBUG] input_tables: {len(req.input_tables)} tables")
    for t in req.input_tables[:3]:
        print(f"  - {t.get('table_name', '???')}: {len(t.get('columns', []))} cols")
    print(f"[DEBUG] output_mapping columns: {len(req.output_mapping.get('columns', []))}")
    if req.output_mapping.get('columns'):
        print(f"  - first: {req.output_mapping['columns'][0].get('name', '???')}")
    print(f"[DEBUG] additional_context: {len(req.additional_context)} chars")

    # === Step1: 方針決定（軽量、Skills無し） ===
    user_message = "アウトプットを実現するための処理方針を決めてください。"
    # additional_contextまたはraw_textからアウトプット原文を取得
    raw_text = req.additional_context or req.output_mapping.get("raw_text", "")
    if raw_text:
        user_message += f"\n\nアウトプット定義の原文:\n{raw_text[:3000]}"

    session["messages"].append({"role": "user", "content": user_message})

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4000,
            system=get_system_prompt_step1(req.input_tables, req.output_mapping),
            messages=session["messages"],
        )
        step1_text = response.content[0].text
        if response.stop_reason == "max_tokens":
            print(f"[WARN] generate Step1 truncated at max_tokens (session={session_id})", flush=True)
    except Exception as e:
        session["messages"].pop()
        return ChatResponse(session_id=session_id, reply=f"API呼び出しエラー: {e}", status="asking")

    session["messages"].append({"role": "assistant", "content": step1_text})

    # Step1の結果を解析
    if "```json" in step1_text:
        try:
            plan_json = json.loads(step1_text.split("```json")[1].split("```")[0].strip())
            if plan_json.get("action") == "plan":
                # Step1成功 → Step2へ自動遷移
                session["plan"] = json.dumps(plan_json, ensure_ascii=False)
                session["step"] = "step2"

                # === Step2: 詳細設計（該当Skillsのみ） ===
                step2_prompt = get_system_prompt_step2(
                    req.input_tables, req.output_mapping, session["plan"]
                )
                session["messages_step2"] = [
                    {"role": "user", "content": "処理方針に基づいて設計書JSONを出力してください。"}
                ]
                try:
                    response2 = client.messages.create(
                        model="claude-sonnet-4-6",
                        max_tokens=16000,
                        system=step2_prompt,
                        messages=session["messages_step2"],
                    )
                    assistant_text = response2.content[0].text
                    session["messages_step2"].append({"role": "assistant", "content": assistant_text})
                except Exception as e:
                    return ChatResponse(session_id=session_id, reply=f"Step2エラー: {e}", status="asking")
        except (json.JSONDecodeError, IndexError):
            assistant_text = step1_text
    else:
        # 技術確認の質問を返している
        assistant_text = step1_text

    # JSON生成チェック
    if "```json" in assistant_text:
        try:
            json_str = assistant_text.split("```json")[1].split("```")[0].strip()
            # JSONが途中で切れてる場合の修復を試みる
            # JSONパース（切れてたら修復を試みる）
            generation_data = _parse_json_with_repair(json_str)
            print(f"[DEBUG] action={generation_data.get('action')}, steps={len(generation_data.get('processing_steps', []))}")

            if generation_data.get("action") == "generate":
                filepath, filename = build_spreadsheet(generation_data)
                session["last_file"] = filepath
                session["last_filename"] = filename

                display_text = assistant_text.split("```json")[0].strip()
                if not display_text:
                    display_text = "手順書を生成しました！以下からダウンロードできます。"

                return ChatResponse(
                    session_id=session_id,
                    reply=display_text,
                    status="done",
                    download_url=f"/api/download/{session_id}",
                )

            elif generation_data.get("action") == "design":
                return _build_from_design_doc(session_id, session, generation_data, req.input_tables)

        except (json.JSONDecodeError, IndexError) as e:
            print(f"[DEBUG] JSON parse error in generate: {e}")
            print(f"[DEBUG] assistant_text[:200]: {assistant_text[:200]}")

    # 質問を返している場合
    return ChatResponse(
        session_id=session_id,
        reply=assistant_text,
        status="asking",
    )


@app.post("/api/chat", response_model=ChatResponse, dependencies=[Depends(verify_token)])
@limiter.limit("20/minute")
async def chat(request: Request, req: ChatRequest):
    """追加質問への回答（モードに応じて分岐）"""
    session = sessions.get(req.session_id)
    if not session:
        raise HTTPException(404, "セッションが見つかりません")
    try:
        return await _chat_body(req, session)
    finally:
        sessions.save(req.session_id, session)


async def _chat_body(req: ChatRequest, session: dict):
    # 施策相談モードの場合
    if session.get("mode") == "consultation":
        return await _handle_consultation_chat(req, session)

    # 手順書モード（既存ロジック）
    session["messages"].append({"role": "user", "content": req.message})

    # Step1を続行（過去の回答を含むコンテキストでAIに判断させる）
    # 過去のQ&Aを要約してプロンプトに追加
    qa_history = []
    for msg in session["messages"]:
        if msg["role"] == "user":
            qa_history.append(msg["content"])
    qa_summary = "\n".join(f"- ユーザー回答: {q}" for q in qa_history[1:])  # 最初のメッセージは除外

    base_prompt = get_system_prompt_step1(session["input_tables"], session["output_mapping"])
    if qa_summary:
        base_prompt += f"\n\n## これまでの技術確認の回答\n{qa_summary}\n\n上記で回答済みの質問は絶対に繰り返さないこと。未確認の項目があれば次の質問をする。全て確認済みならplan JSONを出力する。"

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2000,
            system=base_prompt,
            messages=[{"role": "user", "content": f"回答: {req.message}\n\n未確認の項目があれば次の質問を、全て確認済みならplan JSONを出力してください。"}],
        )
        step1_text = response.content[0].text
    except Exception as e:
        session["messages"].pop()
        print(f"[ERROR] Anthropic API failure: {type(e).__name__}: {e}", flush=True)
        return ChatResponse(session_id=req.session_id, reply=f"エラー: {type(e).__name__}: {e}", status="asking")

    session["messages"].append({"role": "assistant", "content": step1_text})

    # plan JSONが出たら → Step2自動遷移
    if "```json" in step1_text:
        try:
            plan_json = json.loads(step1_text.split("```json")[1].split("```")[0].strip())
            if plan_json.get("action") == "plan":
                session["plan"] = json.dumps(plan_json, ensure_ascii=False)
                session["step"] = "step2"

                step2_prompt = get_system_prompt_step2(
                    session["input_tables"], session["output_mapping"], session["plan"]
                )
                step2_user_msg = f"技術確認完了。回答内容:\n{qa_summary}\n\n設計書JSONを出力してください。"
                session["messages_step2"] = [{"role": "user", "content": step2_user_msg}]
                try:
                    response2 = client.messages.create(
                        model="claude-sonnet-4-6",
                        max_tokens=16000,
                        system=step2_prompt,
                        messages=session["messages_step2"],
                    )
                    assistant_text = response2.content[0].text
                    session["messages_step2"].append({"role": "assistant", "content": assistant_text})
                    # step1_textをassistant_textで上書き（Phase3処理用）
                    step1_text = assistant_text
                except Exception as e:
                    return ChatResponse(session_id=req.session_id, reply=f"Step2エラー: {e}", status="asking")
        except (json.JSONDecodeError, IndexError):
            pass  # plan JSONが出なければ次の質問として返す

    # Step1でplan JSONが出たら → Step2自動遷移
    if "```json" in step1_text:
        try:
            plan_json = json.loads(step1_text.split("```json")[1].split("```")[0].strip())
            if plan_json.get("action") == "plan":
                session["plan"] = json.dumps(plan_json, ensure_ascii=False)
                session["step"] = "step2"

                step2_prompt = get_system_prompt_step2(
                    session["input_tables"], session["output_mapping"], session["plan"]
                )
                session["messages_step2"] = [
                    {"role": "user", "content": "処理方針に基づいて設計書JSONを出力してください。"}
                ]
                try:
                    response2 = client.messages.create(
                        model="claude-sonnet-4-6",
                        max_tokens=16000,
                        system=step2_prompt,
                        messages=session["messages_step2"],
                    )
                    assistant_text = response2.content[0].text
                    session["messages_step2"].append({"role": "assistant", "content": assistant_text})
                except Exception as e:
                    return ChatResponse(session_id=req.session_id, reply=f"Step2エラー: {e}", status="asking")
        except (json.JSONDecodeError, IndexError):
            assistant_text = step1_text
    else:
        assistant_text = step1_text

    # JSON生成チェック（design → Phase3自動実行）
    if "```json" in assistant_text:
        try:
            json_str = assistant_text.split("```json")[1].split("```")[0].strip()
            generation_data = _parse_json_with_repair(json_str)

            if generation_data.get("action") == "generate":
                filepath, filename = build_spreadsheet(generation_data)
                session["last_file"] = filepath
                session["last_filename"] = filename
                display_text = assistant_text.split("```json")[0].strip() or "手順書を生成しました！"
                return ChatResponse(session_id=req.session_id, reply=display_text, status="done", download_url=f"/api/download/{req.session_id}")

            elif generation_data.get("action") == "design":
                # === Phase3: テンプレートエンジンで手順書生成（AI不要） ===
                session["design_doc"] = generation_data
                try:
                    procedure_text = generate_procedure_text(generation_data)
                    session["procedure_text"] = procedure_text

                    steps = generation_data.get("processing_steps", [])
                    proc_rows = []
                    step_num = 1
                    for s in steps:
                        if not isinstance(s, dict):
                            continue
                        step_val = s.get("step", "")
                        sn = str(step_num) if step_val else ""
                        if step_val:
                            step_num += 1
                        op = s.get("operation", "")
                        settings = s.get("settings", {})
                        if not isinstance(settings, dict):
                            settings = {}
                        save_as = s.get("save_as", "")
                        try:
                            template_text = render_step(s)
                        except Exception:
                            template_text = f"『{op}』"
                        param_values = []
                        for v in list(settings.values())[:5]:
                            if isinstance(v, list):
                                parts = [json.dumps(item, ensure_ascii=False) if isinstance(item, dict) else str(item) for item in v]
                                param_values.append(", ".join(parts))
                            elif isinstance(v, dict):
                                param_values.append(json.dumps(v, ensure_ascii=False))
                            else:
                                param_values.append(str(v))
                        while len(param_values) < 5:
                            param_values.append("")
                        proc_rows.append([sn, op, "", template_text, save_as, "", *param_values, template_text])

                    excel_data = {
                        "action": "generate",
                        "title": generation_data.get("summary", "データパレット構築手順書"),
                        "sections": [{"sheet_name": "結合・加工", "title": "手順書",
                            "columns": ["対象作業No", "アイコン", "", "アイコン利用方法", "作成後項目名", "", "対象1", "対象2", "対象3", "対象4", "対象5", "完成形テキスト"],
                            "rows": proc_rows}],
                    }
                    filepath, filename = build_spreadsheet(excel_data)
                    session["last_file"] = filepath
                    session["last_filename"] = filename
                    return ChatResponse(session_id=req.session_id, reply=f"手順書を生成しました！\n\n{procedure_text[:3000]}", status="done", download_url=f"/api/download/{req.session_id}")
                except Exception as e:
                    return ChatResponse(session_id=req.session_id, reply=f"Phase3エラー: {e}", status="asking")
        except (json.JSONDecodeError, IndexError):
            pass

    return ChatResponse(
        session_id=req.session_id,
        reply=assistant_text,
        status="asking",
    )


def _regen_excel(session: dict, session_id: str) -> tuple[str, str]:
    """design_docからExcelを再生成して(filepath, filename)を返す"""
    generation_data = session.get("design_doc")
    if not generation_data:
        raise ValueError("design_docが見つかりません")
    input_tables = session.get("input_tables", [])
    steps = generation_data.get("processing_steps", [])
    proc_rows = []
    step_num = 1
    for s in steps:
        if not isinstance(s, dict):
            continue
        step_val = s.get("step", "")
        sn = str(step_num) if step_val else ""
        if step_val:
            step_num += 1
        op = s.get("operation", "")
        settings = s.get("settings", {})
        if not isinstance(settings, dict):
            settings = {}
        save_as = s.get("save_as", "")
        try:
            template_text = render_step(s)
        except Exception:
            template_text = f"『{op}』"
        param_values = []
        for v in list(settings.values())[:5]:
            if isinstance(v, list):
                parts = [json.dumps(item, ensure_ascii=False) if isinstance(item, dict) else str(item) for item in v]
                param_values.append(", ".join(parts))
            elif isinstance(v, dict):
                param_values.append(json.dumps(v, ensure_ascii=False))
            else:
                param_values.append(str(v))
        while len(param_values) < 5:
            param_values.append("")
        proc_rows.append([sn, op, "", template_text, save_as, "", *param_values, template_text])

    input_names = [t.get("table_name", f"テーブル{i+1}") for i, t in enumerate(input_tables)]
    output_name = next((s.get("save_as", "") for s in reversed(steps) if isinstance(s, dict) and s.get("save_as")), "")
    excel_data = {
        "action": "generate",
        "title": generation_data.get("summary", "データパレット構築手順書"),
        "input_tables": input_names,
        "output_name": output_name,
        "sections": [{"sheet_name": "結合・加工", "title": "手順書",
            "columns": ["対象作業No", "アイコン", "", "アイコン利用方法", "作成後項目名", "", "対象1", "対象2", "対象3", "対象4", "対象5", "完成形テキスト"],
            "rows": proc_rows}],
    }
    filepath, filename = build_spreadsheet(excel_data)
    session["last_file"] = filepath
    session["last_filename"] = filename
    return filepath, filename


@app.get("/api/download/{session_id}", dependencies=[Depends(verify_token)])
async def download(session_id: str):
    session = sessions.get(session_id)
    if not session:
        raise HTTPException(404, "セッションが見つかりません")

    filepath = session.get("last_file")
    filename = session.get("last_filename", f"手順書_{session_id[:8]}.xlsx")

    # ローカルファイルが消えていたらdesign_docから再生成
    if not filepath or not Path(filepath).exists():
        if not session.get("design_doc"):
            raise HTTPException(404, "ファイルが見つかりません")
        try:
            filepath, filename = _regen_excel(session, session_id)
            sessions.save(session_id, session)
        except Exception as e:
            raise HTTPException(500, f"ファイル再生成エラー: {e}")

    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/download-design/{session_id}", dependencies=[Depends(verify_token)])
async def download_design(session_id: str):
    """設計書JSONをダウンロード"""
    session = sessions.get(session_id)
    if not session or "design_file" not in session:
        raise HTTPException(404, "設計書が見つかりません")
    return FileResponse(
        session["design_file"],
        filename=f"設計書_{session_id[:8]}.json",
        media_type="application/json",
    )


# --- ナレッジPDCA ---
class FeedbackRequest(BaseModel):
    session_id: str
    is_correct: bool
    correction: str | None = None

@app.post("/api/feedback", dependencies=[Depends(verify_token)])
async def feedback(req: FeedbackRequest):
    """ユーザーフィードバック → セッションに評価を保存し、可能ならナレッジにも蓄積。

    Phase 1（consultation）/ Phase 3（design）のどちらでも動作する。
    評価は常にセッション本体に書き戻すので CSV エクスポートからも見える。
    """
    session = sessions.get(req.session_id)
    if not session:
        raise HTTPException(404, "セッションが見つかりません")

    # まずセッションに評価を記録（必ず保存）
    session["evaluation"] = "good" if req.is_correct else "bad"
    session["evaluation_correction"] = req.correction or ""
    session["evaluation_at"] = datetime.now().isoformat()
    sessions.save(req.session_id, session)

    design_doc = session.get("design_doc")
    consult = session.get("consultation_result")

    # Phase 3: 設計書フィードバック → ナレッジ蓄積
    if design_doc:
        if req.is_correct:
            _save_knowledge({
                "type": "successful_design",
                "summary": design_doc.get("summary", ""),
                "processing_steps": design_doc.get("processing_steps", []),
                "processing_groups": design_doc.get("processing_groups", []),
            })
            return {"status": "saved", "phase": "design", "message": "ナレッジに保存しました"}
        else:
            _save_knowledge({
                "type": "correction",
                "summary": design_doc.get("summary", ""),
                "correction": req.correction or "",
                "original_steps": design_doc.get("processing_steps", []),
            })
            return {"status": "saved", "phase": "design", "message": "修正内容をナレッジに保存しました"}

    # Phase 1: 施策相談フィードバック → 修正があればナレッジに蓄積
    if consult and not req.is_correct and req.correction:
        _save_knowledge({
            "type": "consultation_correction",
            "strategy_name": consult.get("strategy_name", ""),
            "strategy_summary": consult.get("strategy_summary", ""),
            "correction": req.correction,
        })
        return {"status": "saved", "phase": "consultation", "message": "修正内容をナレッジに保存しました"}

    return {"status": "saved", "phase": "session_only", "message": "評価を記録しました"}


# --- 自由コメント型フィードバック ---
class FeedbackCommentRequest(BaseModel):
    session_id: str | None = None
    phase: str  # 'consultation' | 'organization' | 'design'
    comment: str | None = None  # 単一コメント（後方互換）
    aspects: dict[str, str] | None = None  # {aspect_key: text} — 観点別コメント
    card_context: dict | None = None


def _build_feedback_output(session: dict | None, phase: str) -> tuple[dict | None, str]:
    """セッションとフェーズから、FB保存用の (output_dict, output_text) を作る。

    output_text: 分析者がCSV内で直読できるテキスト
    output_dict: card_context に同梱する生データ
    """
    if not session:
        return None, ""

    if phase == "consultation":
        cr = session.get("consultation_result") or {}
        if not cr:
            return None, ""
        req = cr.get("requirements") or {}
        lines = [
            f"戦略名: {cr.get('strategy_name', '')}",
            f"概要: {cr.get('strategy_summary', '')}",
            f"① 誰に: {req.get('who', '')}",
            f"② 何を: {req.get('what', '')}",
            f"③ いつ: {req.get('when', '')}",
            f"④ 除外: {req.get('exclude', '')}",
        ]
        out_cols = cr.get("output_columns") or []
        if out_cols:
            lines.append("アウトプット項目:")
            for c in out_cols:
                if isinstance(c, dict):
                    src = c.get("source_column") or c.get("definition", "")
                    lines.append(f"  - {c.get('name', '')}: {src}")
                else:
                    lines.append(f"  - {c}")
        in_tabs = cr.get("input_tables") or []
        if in_tabs:
            names = [t.get("table_name", "") if isinstance(t, dict) else str(t) for t in in_tabs]
            lines.append(f"インプットテーブル: {', '.join(names)}")
        return cr, "\n".join(lines)

    if phase == "organization":
        cr = session.get("consultation_result") or {}
        input_tables = session.get("input_tables") or []
        output_mapping = session.get("output_mapping") or {}
        lines: list[str] = []
        if input_tables:
            lines.append("【インプットテーブル】")
            for t in input_tables:
                if not isinstance(t, dict):
                    continue
                tn = t.get("table_name", "")
                cols = t.get("columns") or []
                col_names = [c.get("name", "") if isinstance(c, dict) else str(c) for c in cols]
                lines.append(f"- {tn}: {', '.join(col_names)}")
        if output_mapping.get("columns"):
            lines.append("【アウトプットマッピング】")
            for col in output_mapping["columns"]:
                if not isinstance(col, dict):
                    continue
                src_tbl = col.get("source_table", "") or ""
                src_col = col.get("source_column", "") or "（加工生成）"
                lines.append(f"- {col.get('name', '')}: {col.get('definition', '')} ← {src_tbl}.{src_col}")
        if not lines and cr:
            lines.append(f"戦略名: {cr.get('strategy_name', '')}")
            lines.append(f"概要: {cr.get('strategy_summary', '')}")
        snapshot = {
            "consultation_result": cr,
            "input_tables": input_tables,
            "output_mapping": output_mapping,
        }
        return snapshot, "\n".join(lines)

    if phase == "design":
        dd = session.get("design_doc") or {}
        if not dd:
            return None, ""
        lines = [f"概要: {dd.get('summary', '')}"]
        steps = dd.get("processing_steps") or []
        if steps:
            lines.append("【処理ステップ】")
            for s in steps:
                if isinstance(s, dict):
                    n = s.get("step", "")
                    op = s.get("operation", "")
                    result = s.get("result", "") or s.get("save_as", "")
                    lines.append(f"  {n}. {op}: {result}".rstrip(": "))
                else:
                    lines.append(f"  - {s}")
        notes = dd.get("special_notes") or []
        if notes:
            lines.append("【注意事項】")
            for n in notes:
                lines.append(f"  - {n}")
        return dd, "\n".join(lines)

    return None, ""


@app.post("/api/feedback/comment", dependencies=[Depends(verify_token)])
async def feedback_comment(req: FeedbackCommentRequest):
    """各フェーズのカードから自由テキストFBを受け取り、feedback_log に永続化する。

    aspects が指定された場合は観点ごとに1行ずつ insert する（空文字は無視）。
    aspects 未指定で comment のみあれば、aspect=null で1行 insert する。
    保存時、session_id+phase からその時点のアウトプットを引いて output_text に格納する。
    """
    phase = (req.phase or "").strip() or "unknown"

    # セッションからアウトプットスナップショットを取り出す
    session = sessions.get(req.session_id) if req.session_id else None
    output_dict, output_text = _build_feedback_output(session, phase)

    # card_context: フロント送信分 + アウトプット生データをマージ
    ctx: dict = dict(req.card_context or {})
    if output_dict:
        ctx["output"] = output_dict
    ctx_json = Json(ctx) if ctx else None

    # 挿入対象を (aspect, text) のリストにまとめる
    entries: list[tuple[str | None, str]] = []
    if req.aspects:
        for aspect, text in req.aspects.items():
            t = (text or "").strip()
            if t:
                entries.append((aspect, t))
    if not entries and req.comment:
        c = req.comment.strip()
        if c:
            entries.append((None, c))

    if not entries:
        raise HTTPException(400, "コメントが空です")

    inserted: list[dict] = []
    with _db_conn() as conn:
        if conn is None:
            raise HTTPException(503, "DB未設定")
        with conn.cursor() as cur:
            for aspect, text in entries:
                cur.execute(
                    "INSERT INTO feedback_log "
                    "(session_id, phase, aspect, comment, output_text, card_context) "
                    "VALUES (%s, %s, %s, %s, %s, %s) RETURNING id, created_at",
                    (req.session_id, phase, aspect, text, output_text or None, ctx_json),
                )
                row = cur.fetchone()
                inserted.append({
                    "id": row[0] if row else None,
                    "aspect": aspect,
                    "created_at": row[1].isoformat() if row and row[1] else None,
                })

    return {"status": "saved", "count": len(inserted), "entries": inserted}


@app.get("/api/feedback/export", dependencies=[Depends(verify_token)])
async def export_feedback_csv():
    """feedback_log を CSV で出力（Excel互換のUTF-8 BOM付き）。"""
    header = [
        "id", "created_at", "session_id", "phase", "aspect", "comment",
        "output_text", "card_context",
    ]
    rows: list[list[str]] = [header]

    with _db_conn() as conn:
        if conn is None:
            raise HTTPException(503, "DB未設定")
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                "SELECT id, created_at, session_id, phase, aspect, comment, "
                "output_text, card_context "
                "FROM feedback_log ORDER BY created_at DESC"
            )
            db_rows = cur.fetchall()

    for r in db_rows:
        ctx = r.get("card_context")
        ctx_str = json.dumps(ctx, ensure_ascii=False) if ctx else ""
        rows.append([
            str(r.get("id", "")),
            r["created_at"].isoformat() if r.get("created_at") else "",
            r.get("session_id") or "",
            r.get("phase") or "",
            r.get("aspect") or "",
            r.get("comment") or "",
            r.get("output_text") or "",
            ctx_str,
        ])

    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_ALL)
    for row in rows:
        writer.writerow(row)
    body = "﻿" + buf.getvalue()
    filename = f"feedback_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        content=body.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- 施策相談エンドポイント ---

@app.post("/api/consultation/start", response_model=ChatResponse, dependencies=[Depends(verify_token)])
@limiter.limit("10/minute")
async def consultation_start(request: Request, req: ConsultationStartRequest):
    """施策相談セッションを開始"""
    session_id = str(uuid.uuid4())
    session = {
        "mode": "consultation",
        "messages": [],
        "industry": req.industry,
        "consultation_result": None,
        "question_count": 0,
        "created_at": _time.time(),
    }
    try:
        return await _consultation_start_body(req, session_id, session)
    finally:
        sessions.save(session_id, session)


async def _consultation_start_body(req: ConsultationStartRequest, session_id: str, session: dict):
    industries = _load_industries()
    system_prompt = get_consultation_system_prompt(req.industry)
    user_message = req.message
    if req.industry and req.industry in industries:
        user_message += f"\n\n（業界: {industries[req.industry].get('label', req.industry)}）"

    session["messages"].append({"role": "user", "content": user_message})

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=16000,
            system=system_prompt,
            messages=session["messages"],
        )
        reply_text = response.content[0].text
        if response.stop_reason == "max_tokens":
            print(f"[WARN] consultation reply truncated at max_tokens (start endpoint)", flush=True)
    except Exception as e:
        session["messages"].pop()
        print(f"[ERROR] Anthropic API failure: {type(e).__name__}: {e}", flush=True)
        return ChatResponse(session_id=session_id, reply=f"エラー: {type(e).__name__}: {e}", status="asking")

    session["messages"].append({"role": "assistant", "content": reply_text})
    session["question_count"] = session.get("question_count", 0) + 1

    result = _check_consultation_result(reply_text, session)
    if result:
        return ChatResponse(
            session_id=session_id,
            reply=_CONSULT_RESULT_BLOCK_RE.sub("", reply_text).strip() or "施策設計が完了しました！",
            status="consultation_complete",
            consultation_result=result,
        )

    return ChatResponse(session_id=session_id, reply=reply_text, status="asking")


@app.post("/api/consultation/apply", response_model=ChatResponse, dependencies=[Depends(verify_token)])
async def consultation_apply(req: ConsultationApplyRequest):
    """施策相談の結果を手順書パイプラインに橋渡し（互換用、後方互換のため残存）"""
    session = sessions.get(req.session_id)
    if not session:
        raise HTTPException(404, "セッションが見つかりません")

    consultation_result = session.get("consultation_result")
    if not consultation_result:
        raise HTTPException(400, "施策相談が完了していません")

    input_tables = req.input_tables or consultation_result["input_tables"]
    output_mapping = req.output_mapping or consultation_result["output_mapping"]

    _save_knowledge({
        "type": "consultation",
        "strategy_name": consultation_result.get("strategy_name", ""),
        "strategy_summary": consultation_result.get("strategy_summary", ""),
        "output_columns": [c.get("name", "") for c in output_mapping.get("columns", [])],
        "input_table_names": [t.get("table_name", "") for t in input_tables],
    })

    new_session_id = str(uuid.uuid4())
    generate_req = GenerateRequest(
        session_id=new_session_id,
        input_tables=input_tables,
        output_mapping=output_mapping,
        additional_context="",
    )
    result = await _generate_core(generate_req)
    result.session_id = new_session_id
    return result


# ========== テーブル定義整理フェーズ (Phase2) ==========

_ORG_BLOCK_RE = re.compile(r"```\s*(?:json|JSON)?\s*(\{.*?\})\s*```", re.DOTALL)


def _build_org_payload(session_id: str, text: str, session: dict) -> dict:
    """テーブル定義整理エージェントの返信をパース。Phase1と同じく
    フェンス変形・複数ブロック・raw fallback に耐え、抽出失敗時は [WARN] preview を残す。
    session を mutate する副作用あり (review_state / output_mapping / finalized)。"""
    candidates: list[str] = [m.group(1) for m in _ORG_BLOCK_RE.finditer(text)]

    if not candidates and ('"organization_review"' in text or '"organization_complete"' in text):
        start = text.find("{")
        end = text.rfind("}")
        if start != -1 and end > start:
            candidates.append(text[start:end + 1])

    payload = None
    for js in candidates:
        try:
            data = json.loads(js)
        except json.JSONDecodeError:
            continue
        if isinstance(data, dict) and data.get("action") in ("organization_review", "organization_complete"):
            payload = data
            break

    if payload is None:
        if "organization_review" in text or "organization_complete" in text:
            print(
                f"[WARN] organization payload mentioned but not parseable. len={len(text)} head={text[:300]!r} tail={text[-300:]!r}",
                flush=True,
            )
        return {"session_id": session_id, "reply": text, "status": "asking"}

    action = payload.get("action")
    display_text = _ORG_BLOCK_RE.sub("", text).strip() or "マッピング候補を作成しました。"

    if action == "organization_review":
        session["review_state"] = payload
        return {
            "session_id": session_id,
            "reply": display_text,
            "status": "organization_review",
            "consultation_result": payload,
        }

    # organization_complete
    session["input_tables"] = payload.get("input_tables", session.get("input_tables", []))
    session["output_mapping"] = payload.get("output_mapping")
    session["finalized"] = True
    return {
        "session_id": session_id,
        "reply": display_text or "マッピングが確定しました。",
        "status": "organization_complete",
        "consultation_result": payload,
    }


def _sse_pack(payload: dict) -> str:
    """1つのSSEイベントをエンコード。data行は1行に収める (JSON内の改行は\\nにエスケープ)。"""
    return f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"


async def _stream_org_mapping(
    session: dict,
    session_id: str,
    *,
    initial_user_msg: str | None = None,
    appended_user_msg: str | None = None,
):
    """SSE async generator. type=token (delta), type=final (parsed payload), type=error.
    完了時 (success/error両方) に sessions.save を呼ぶ。"""
    rollback_pop = False
    if initial_user_msg is not None:
        session["messages"] = [{"role": "user", "content": initial_user_msg}]
    elif appended_user_msg is not None:
        session["messages"].append({"role": "user", "content": appended_user_msg})
        rollback_pop = True

    chunks: list[str] = []
    try:
        async with async_client.messages.stream(
            model="claude-sonnet-4-6",
            max_tokens=16000,
            system=get_organization_system_prompt(session),
            messages=session["messages"],
        ) as stream:
            async for delta in stream.text_stream:
                chunks.append(delta)
                yield _sse_pack({"type": "token", "text": delta})
            final_message = await stream.get_final_message()
            if final_message.stop_reason == "max_tokens":
                print("[WARN] organization reply truncated at max_tokens", flush=True)
    except Exception as e:
        if rollback_pop and session["messages"] and session["messages"][-1].get("role") == "user":
            session["messages"].pop()
        sessions.save(session_id, session)
        print(f"[ERROR] Anthropic stream failure: {type(e).__name__}: {e}", flush=True)
        yield _sse_pack({"type": "error", "error": f"{type(e).__name__}: {e}"})
        return

    reply_text = "".join(chunks)
    session["messages"].append({"role": "assistant", "content": reply_text})

    payload = _build_org_payload(session_id, reply_text, session)
    sessions.save(session_id, session)

    yield _sse_pack({"type": "final", **payload})


_SSE_HEADERS = {"X-Accel-Buffering": "no", "Cache-Control": "no-cache"}


@app.post("/api/organization/start", dependencies=[Depends(verify_token)])
@limiter.limit("10/minute")
async def organization_start(request: Request, req: OrganizationStartRequest):
    """Phase1完了後にテーブル定義整理フェーズを開始 (SSEストリーミング)"""
    consult = sessions.get(req.consultation_session_id)
    if not consult:
        raise HTTPException(404, "施策相談セッションが見つかりません")
    if not consult.get("consultation_result"):
        raise HTTPException(400, "施策相談が完了していません")

    org_id = str(uuid.uuid4())
    org_session = {
        "mode": "organization",
        "consultation_session_id": req.consultation_session_id,
        "consultation_result": consult["consultation_result"],
        "input_tables": req.input_tables,
        "output_mapping": None,
        "messages": [],
        "review_state": None,
        "finalized": False,
        "created_at": _time.time(),
    }

    user_msg = "上記の実テーブルに基づいて、施策要件をマッピングしてください。"
    if req.additional_hint:
        user_msg += f"\n\n補足: {req.additional_hint}"

    return StreamingResponse(
        _stream_org_mapping(org_session, org_id, initial_user_msg=user_msg),
        media_type="text/event-stream",
        headers=_SSE_HEADERS,
    )


@app.post("/api/organization/chat", dependencies=[Depends(verify_token)])
@limiter.limit("20/minute")
async def organization_chat(request: Request, req: OrganizationChatRequest):
    """テーブル定義整理フェーズのチャット (SSEストリーミング)"""
    session = sessions.get(req.session_id)
    if not session or session.get("mode") != "organization":
        raise HTTPException(404, "organizationセッションが見つかりません")

    return StreamingResponse(
        _stream_org_mapping(session, req.session_id, appended_user_msg=req.message),
        media_type="text/event-stream",
        headers=_SSE_HEADERS,
    )


@app.post("/api/organization/update-tables", dependencies=[Depends(verify_token)])
async def organization_update_tables(req: OrganizationUpdateTablesRequest):
    """実テーブルを差し替えて再マッピング (SSEストリーミング)"""
    session = sessions.get(req.session_id)
    if not session or session.get("mode") != "organization":
        raise HTTPException(404, "organizationセッションが見つかりません")

    session["input_tables"] = req.input_tables
    session["output_mapping"] = None
    session["review_state"] = None
    session["finalized"] = False

    user_msg = "上記の実テーブルに基づいて、施策要件をマッピングしてください。"
    return StreamingResponse(
        _stream_org_mapping(session, req.session_id, initial_user_msg=user_msg),
        media_type="text/event-stream",
        headers=_SSE_HEADERS,
    )


@app.get("/api/organization/{session_id}/csv", dependencies=[Depends(verify_token)])
async def organization_csv(session_id: str):
    """Phase2完了後のアウトプット定義をCSVでダウンロード"""
    session = sessions.get(session_id)
    if not session:
        # DBからも試みる
        with _db_conn() as conn:
            if conn is not None:
                with conn.cursor() as cur:
                    cur.execute("SELECT data FROM sessions WHERE id = %s", (session_id,))
                    row = cur.fetchone()
                    if row:
                        session = row[0]
    if not session:
        raise HTTPException(404, "セッションが見つかりません")

    output_mapping = session.get("output_mapping") or {}
    columns = output_mapping.get("columns", [])
    if not columns:
        raise HTTPException(404, "アウトプット定義がありません")

    PURPOSE_LABEL = {"segment": "セグメント", "personalize": "差し込み"}
    header = "カラム名,用途,定義,ソーステーブル,ソースカラム"

    def escape_csv(val: str) -> str:
        val = str(val or "")
        if any(c in val for c in (',', '"', '\n')):
            val = '"' + val.replace('"', '""') + '"'
        return val

    lines = [header]
    for col in columns:
        purpose = PURPOSE_LABEL.get(col.get("purpose", ""), col.get("purpose", ""))
        lines.append(",".join(escape_csv(v) for v in [
            col.get("name", ""),
            purpose,
            col.get("definition", ""),
            col.get("source_table", ""),
            col.get("source_column", ""),
        ]))

    body = "﻿" + "\n".join(lines)  # BOM付きUTF-8（Excel対応）
    strategy = (session.get("consultation_result") or {}).get("strategy_name", "output_mapping")
    from urllib.parse import quote
    encoded = quote(strategy[:40] + ".csv", safe="")
    return Response(
        content=body.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=\"output_mapping.csv\"; filename*=UTF-8''{encoded}"},
    )


@app.post("/api/organization/finalize", response_model=ChatResponse, dependencies=[Depends(verify_token)])
@limiter.limit("10/minute")
async def organization_finalize(request: Request, req: OrganizationFinalizeRequest):
    """整理結果を既存の/api/generateに引き渡して手順書生成"""
    session = sessions.get(req.session_id)
    if not session or session.get("mode") != "organization":
        raise HTTPException(404, "organizationセッションが見つかりません")

    if not session.get("output_mapping"):
        raise HTTPException(400, "マッピングが未確定です")

    input_tables = session["input_tables"]
    output_mapping = session["output_mapping"]
    cr = session.get("consultation_result") or {}

    _save_knowledge({
        "type": "organization",
        "strategy_name": cr.get("strategy_name", ""),
        "strategy_summary": cr.get("strategy_summary", ""),
        "output_columns": [c.get("name", "") for c in output_mapping.get("columns", [])],
        "input_table_names": [t.get("table_name", "") for t in input_tables],
    })

    new_session_id = str(uuid.uuid4())
    generate_req = GenerateRequest(
        session_id=new_session_id,
        input_tables=input_tables,
        output_mapping=output_mapping,
        additional_context="",
    )
    result = await _generate_core(generate_req)
    result.session_id = new_session_id
    return result


async def _handle_consultation_chat(req: ChatRequest, session: dict) -> ChatResponse:
    """施策相談モードのチャットハンドラ"""
    session["messages"].append({"role": "user", "content": req.message})
    session["question_count"] = session.get("question_count", 0) + 1

    industry = session.get("industry")
    if not industry:
        industries = _load_industries()
        for key, ind in industries.items():
            if ind.get("label", "") in req.message:
                session["industry"] = key
                industry = key
                break

    system_prompt = get_consultation_system_prompt(industry)

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=16000,
            system=system_prompt,
            messages=session["messages"],
        )
        reply_text = response.content[0].text
        if response.stop_reason == "max_tokens":
            print(f"[WARN] consultation reply truncated at max_tokens (chat endpoint)", flush=True)
    except Exception as e:
        session["messages"].pop()
        print(f"[ERROR] Anthropic API failure: {type(e).__name__}: {e}", flush=True)
        return ChatResponse(session_id=req.session_id, reply=f"エラー: {type(e).__name__}: {e}", status="asking")

    session["messages"].append({"role": "assistant", "content": reply_text})

    result = _check_consultation_result(reply_text, session)
    if result:
        return ChatResponse(
            session_id=req.session_id,
            reply=_CONSULT_RESULT_BLOCK_RE.sub("", reply_text).strip() or "施策設計が完了しました！",
            status="consultation_complete",
            consultation_result=result,
        )

    return ChatResponse(session_id=req.session_id, reply=reply_text, status="asking")


_CONSULT_RESULT_BLOCK_RE = re.compile(r"```\s*(?:json|JSON)?\s*(\{.*?\})\s*```", re.DOTALL)


def _check_consultation_result(text: str, session: dict) -> dict | None:
    """AIレスポンスからconsultation_result JSONを抽出。
    フェンス言語タグの大小・空白ゆらぎ、複数ブロック、フェンス無しのraw JSONにも耐える。"""
    candidates: list[str] = [m.group(1) for m in _CONSULT_RESULT_BLOCK_RE.finditer(text)]

    if not candidates and '"consultation_result"' in text:
        # フェンス無しでJSONをそのまま吐いたケース: 一番外側の {...} をざっくり拾う
        start = text.find("{")
        end = text.rfind("}")
        if start != -1 and end > start:
            candidates.append(text[start:end + 1])

    for js in candidates:
        try:
            data = json.loads(js)
        except json.JSONDecodeError:
            continue
        if isinstance(data, dict) and data.get("action") == "consultation_result":
            session["consultation_result"] = data
            return data

    if "consultation_result" in text:
        # action=consultation_result を意図したが取り出せなかった: 原因調査用に先頭400字を残す
        print(
            f"[WARN] consultation_result mentioned but not parseable. preview={text[:400]!r}",
            flush=True,
        )
    return None


# ========== プリセット/テンプレート管理API ==========

@app.get("/api/strategy-templates", dependencies=[Depends(verify_token)])
async def list_strategy_templates():
    """施策テンプレート一覧"""
    return {"entries": _load_strategy_templates()}


@app.post("/api/strategy-templates", dependencies=[Depends(verify_token)])
async def create_strategy_template(entry: dict):
    """施策テンプレート追加"""
    if not entry.get("id"):
        entry["id"] = str(uuid.uuid4())[:8]
    _upsert("strategy_templates", entry, _STRATEGY_COLS)
    return {"status": "created", "id": entry["id"]}


@app.put("/api/strategy-templates/{tpl_id}", dependencies=[Depends(verify_token)])
async def update_strategy_template(tpl_id: str, update: dict):
    """施策テンプレート更新"""
    update.pop("id", None)
    update["updated_at"] = datetime.now().isoformat()
    _update_row("strategy_templates", tpl_id, update, _STRATEGY_COLS)
    return {"status": "updated"}


@app.delete("/api/strategy-templates/{tpl_id}", dependencies=[Depends(verify_token)])
async def delete_strategy_template(tpl_id: str):
    """施策テンプレート削除"""
    _delete_row("strategy_templates", tpl_id)
    return {"status": "deleted"}


@app.get("/api/industries/deleted", dependencies=[Depends(verify_token)])
async def list_deleted_industries():
    """ソフトデリート済みの業界プリセット一覧（30日以内）"""
    try:
        with _db_conn() as conn:
            if conn is None:
                return {}
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    "SELECT * FROM industries WHERE deleted_at IS NOT NULL "
                    "AND deleted_at > NOW() - INTERVAL '30 days' ORDER BY deleted_at DESC"
                )
                rows = cur.fetchall()
                return {row["id"]: dict(row) for row in rows}
    except Exception:
        return {}


@app.post("/api/industries", dependencies=[Depends(verify_token)])
async def create_industry(entry: dict):
    """業界プリセット追加"""
    if not entry.get("id"):
        entry["id"] = str(uuid.uuid4())[:8]
    _upsert("industries", entry, _INDUSTRY_COLS)
    return {"status": "created", "id": entry["id"]}


@app.put("/api/industries/{ind_id}", dependencies=[Depends(verify_token)])
async def update_industry(ind_id: str, update: dict):
    """業界プリセット更新"""
    update.pop("id", None)
    update["updated_at"] = datetime.now().isoformat()
    _update_row("industries", ind_id, update, _INDUSTRY_COLS)
    return {"status": "updated"}


@app.post("/api/industries/{ind_id}/restore", dependencies=[Depends(verify_token)])
async def restore_industry(ind_id: str):
    """ソフトデリート済みプリセットを復元"""
    with _db_conn() as conn:
        if conn is None:
            raise RuntimeError("DB未設定")
        with conn.cursor() as cur:
            cur.execute("UPDATE industries SET deleted_at = NULL WHERE id = %s", (ind_id,))
    return {"status": "restored"}


@app.delete("/api/industries/{ind_id}", dependencies=[Depends(verify_token)])
async def delete_industry(ind_id: str):
    """業界プリセットをソフトデリート（30日後に自動失効）"""
    with _db_conn() as conn:
        if conn is None:
            raise RuntimeError("DB未設定")
        with conn.cursor() as cur:
            cur.execute("UPDATE industries SET deleted_at = NOW() WHERE id = %s", (ind_id,))
    return {"status": "deleted"}


# --- フロントエンド配信 ---
FRONTEND_PATH = BASE_DIR / "frontend"

# ========== ナレッジ管理API ==========

@app.get("/api/knowledge", dependencies=[Depends(verify_token)])
async def get_knowledge():
    """ナレッジ一覧を返す"""
    kb = _load_knowledge_base()
    return {"entries": kb, "total": len(kb)}


@app.get("/api/knowledge/{entry_id}", dependencies=[Depends(verify_token)])
async def get_knowledge_entry(entry_id: str):
    """ナレッジ1件を返す"""
    kb = _load_knowledge_base()
    for entry in kb:
        if entry.get("id") == entry_id:
            return entry
    raise HTTPException(404, "ナレッジが見つかりません")


@app.put("/api/knowledge/{entry_id}", dependencies=[Depends(verify_token)])
async def update_knowledge_entry(entry_id: str, update: dict):
    """ナレッジ1件を更新"""
    result = _update_knowledge(entry_id, update)
    if result is None:
        raise HTTPException(404, "ナレッジが見つかりません")
    return {"status": "updated", "entry": result}


@app.delete("/api/knowledge/{entry_id}", dependencies=[Depends(verify_token)])
async def delete_knowledge_entry(entry_id: str):
    """ナレッジ1件を削除"""
    if not _delete_knowledge(entry_id):
        raise HTTPException(404, "ナレッジが見つかりません")
    return {"status": "deleted"}


@app.get("/api/usage/export", dependencies=[Depends(verify_token)])
async def export_usage_xlsx():
    """利用状況集計をExcel(.xlsx)で出力。要約/日別/Mode別/FB集計の4シート。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    P1_DONE_SQL = "(mode = 'consultation' AND data->'consultation_result' IS NOT NULL)"
    P2_DONE_SQL = "(mode = 'organization' AND data->'output_mapping' IS NOT NULL)"
    P3_DONE_SQL = "(mode = 'procedure' AND data->>'last_file' IS NOT NULL)"
    DONE_ANY = f"({P1_DONE_SQL} OR {P2_DONE_SQL} OR {P3_DONE_SQL})"

    with _db_conn() as conn:
        if conn is None:
            raise HTTPException(503, "DB未設定")
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # --- Mode別集計（要約・Mode別 共通の素データ）---
            cur.execute(f"""
                SELECT
                    mode,
                    COUNT(*) AS total,
                    SUM(CASE WHEN {DONE_ANY} THEN 1 ELSE 0 END) AS done,
                    AVG(EXTRACT(EPOCH FROM (updated_at - created_at))/60.0) AS avg_min,
                    percentile_cont(0.5) WITHIN GROUP (
                        ORDER BY EXTRACT(EPOCH FROM (updated_at - created_at))/60.0
                    ) AS median_min,
                    MAX(EXTRACT(EPOCH FROM (updated_at - created_at))/60.0) AS max_min,
                    AVG(jsonb_array_length(COALESCE(data->'messages', '[]'::jsonb))) AS avg_msgs,
                    SUM(CASE WHEN data->>'evaluation' = 'good' THEN 1 ELSE 0 END) AS eval_good,
                    SUM(CASE WHEN data->>'evaluation' = 'bad'  THEN 1 ELSE 0 END) AS eval_bad
                FROM sessions
                GROUP BY mode
                ORDER BY mode
            """)
            mode_rows = cur.fetchall()

            # --- 全期間トータル ---
            cur.execute(f"""
                SELECT
                    COUNT(*) AS total,
                    SUM(CASE WHEN {DONE_ANY} THEN 1 ELSE 0 END) AS done,
                    MIN(created_at) AS first_at,
                    MAX(created_at) AS last_at
                FROM sessions
            """)
            overall = cur.fetchone() or {}

            # --- 日別集計（直近90日, JST）---
            cur.execute(f"""
                SELECT
                    DATE(created_at AT TIME ZONE 'Asia/Tokyo') AS day,
                    COUNT(*) FILTER (WHERE mode = 'consultation') AS p1_total,
                    COUNT(*) FILTER (WHERE {P1_DONE_SQL}) AS p1_done,
                    COUNT(*) FILTER (WHERE mode = 'organization') AS p2_total,
                    COUNT(*) FILTER (WHERE {P2_DONE_SQL}) AS p2_done,
                    COUNT(*) FILTER (WHERE mode = 'procedure') AS p3_total,
                    COUNT(*) FILTER (WHERE {P3_DONE_SQL}) AS p3_done,
                    AVG(EXTRACT(EPOCH FROM (updated_at - created_at))/60.0) AS avg_min
                FROM sessions
                WHERE created_at >= NOW() - INTERVAL '90 days'
                GROUP BY day
                ORDER BY day DESC
            """)
            daily_rows = cur.fetchall()

            # --- FB集計 (feedback_log: aspect別コメント件数) ---
            cur.execute("""
                SELECT
                    COALESCE(phase, '') AS phase,
                    COALESCE(aspect, '') AS aspect,
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE COALESCE(TRIM(comment), '') <> '') AS non_empty
                FROM feedback_log
                GROUP BY phase, aspect
                ORDER BY phase, aspect
            """)
            fb_rows = cur.fetchall()

    def _r(v, ndigits=1):
        if v is None:
            return ""
        try:
            return round(float(v), ndigits)
        except (TypeError, ValueError):
            return ""

    def _rate(num, denom):
        try:
            n = float(num or 0)
            d = float(denom or 0)
            if d <= 0:
                return ""
            return round(n / d * 100, 1)
        except (TypeError, ValueError):
            return ""

    # --- Workbook作成 ---
    wb = Workbook()
    HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    HEADER_FONT = Font(bold=True)
    WRAP = Alignment(wrap_text=True, vertical="top")

    def _write_header(ws, headers: list[str]):
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = WRAP

    def _autofit(ws, widths: list[int]):
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    MODE_LABEL = {
        "consultation": "Phase1 (施策相談)",
        "organization": "Phase2 (テーブル整理)",
        "procedure":    "Phase3 (手順書生成)",
    }

    # ----- Sheet 1: 要約 -----
    ws = wb.active
    ws.title = "要約"
    _write_header(ws, ["項目", "値"])
    total_all = overall.get("total") or 0
    done_all = overall.get("done") or 0
    first_at = overall.get("first_at")
    last_at = overall.get("last_at")
    summary_pairs = [
        ("集計期間 (最古)", first_at.isoformat() if first_at else ""),
        ("集計期間 (最新)", last_at.isoformat() if last_at else ""),
        ("総セッション数", total_all),
        ("完了セッション数", done_all),
        ("完了率(%)", _rate(done_all, total_all)),
        ("", ""),
        ("--- Mode別 ---", ""),
    ]
    for mr in mode_rows:
        label = MODE_LABEL.get(mr["mode"] or "", mr["mode"] or "(不明)")
        summary_pairs.append((f"{label} セッション数", mr["total"]))
        summary_pairs.append((f"{label} 完了数", mr["done"]))
        summary_pairs.append((f"{label} 完了率(%)", _rate(mr["done"], mr["total"])))
        summary_pairs.append((f"{label} 平均所要時間(分)", _r(mr["avg_min"])))
    summary_pairs.append(("", ""))
    summary_pairs.append(("--- 評価 (✓/✗) ---", ""))
    tot_good = sum((mr["eval_good"] or 0) for mr in mode_rows)
    tot_bad = sum((mr["eval_bad"] or 0) for mr in mode_rows)
    summary_pairs.append(("✓ 正しい 件数", tot_good))
    summary_pairs.append(("✗ 修正が必要 件数", tot_bad))
    summary_pairs.append(("✓ 比率(%)", _rate(tot_good, tot_good + tot_bad)))

    for i, (k, v) in enumerate(summary_pairs, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    _autofit(ws, [32, 30])

    # ----- Sheet 2: 日別 -----
    ws2 = wb.create_sheet("日別")
    _write_header(ws2, [
        "日付", "P1 開始", "P1 完了", "P1 完了率(%)",
        "P2 開始", "P2 完了", "P2 完了率(%)",
        "P3 開始", "P3 完了", "P3 完了率(%)",
        "平均所要時間(分)",
    ])
    for i, dr in enumerate(daily_rows, start=2):
        ws2.cell(row=i, column=1, value=dr["day"].isoformat() if dr.get("day") else "")
        ws2.cell(row=i, column=2, value=dr["p1_total"])
        ws2.cell(row=i, column=3, value=dr["p1_done"])
        ws2.cell(row=i, column=4, value=_rate(dr["p1_done"], dr["p1_total"]))
        ws2.cell(row=i, column=5, value=dr["p2_total"])
        ws2.cell(row=i, column=6, value=dr["p2_done"])
        ws2.cell(row=i, column=7, value=_rate(dr["p2_done"], dr["p2_total"]))
        ws2.cell(row=i, column=8, value=dr["p3_total"])
        ws2.cell(row=i, column=9, value=dr["p3_done"])
        ws2.cell(row=i, column=10, value=_rate(dr["p3_done"], dr["p3_total"]))
        ws2.cell(row=i, column=11, value=_r(dr["avg_min"]))
    _autofit(ws2, [12, 8, 8, 12, 8, 8, 12, 8, 8, 12, 16])

    # ----- Sheet 3: Mode別 -----
    ws3 = wb.create_sheet("Mode別")
    _write_header(ws3, [
        "Mode", "開始数", "完了数", "完了率(%)",
        "平均所要時間(分)", "中央値(分)", "最大(分)", "平均メッセージ数",
        "✓ 件数", "✗ 件数",
    ])
    for i, mr in enumerate(mode_rows, start=2):
        ws3.cell(row=i, column=1, value=MODE_LABEL.get(mr["mode"] or "", mr["mode"] or "(不明)"))
        ws3.cell(row=i, column=2, value=mr["total"])
        ws3.cell(row=i, column=3, value=mr["done"])
        ws3.cell(row=i, column=4, value=_rate(mr["done"], mr["total"]))
        ws3.cell(row=i, column=5, value=_r(mr["avg_min"]))
        ws3.cell(row=i, column=6, value=_r(mr["median_min"]))
        ws3.cell(row=i, column=7, value=_r(mr["max_min"]))
        ws3.cell(row=i, column=8, value=_r(mr["avg_msgs"]))
        ws3.cell(row=i, column=9, value=mr["eval_good"])
        ws3.cell(row=i, column=10, value=mr["eval_bad"])
    _autofit(ws3, [22, 10, 10, 12, 16, 14, 12, 16, 10, 10])

    # ----- Sheet 4: FB集計 -----
    ws4 = wb.create_sheet("FB集計")
    _write_header(ws4, ["Phase", "Aspect", "コメント総件数", "記入あり", "記入率(%)"])
    for i, fr in enumerate(fb_rows, start=2):
        ws4.cell(row=i, column=1, value=fr["phase"] or "(不明)")
        ws4.cell(row=i, column=2, value=fr["aspect"] or "(不明)")
        ws4.cell(row=i, column=3, value=fr["total"])
        ws4.cell(row=i, column=4, value=fr["non_empty"])
        ws4.cell(row=i, column=5, value=_rate(fr["non_empty"], fr["total"]))
    _autofit(ws4, [16, 22, 14, 12, 12])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"usage_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/sessions/export", dependencies=[Depends(verify_token)])
async def export_sessions_csv():
    """全セッションをCSVで出力（現場メンバーの精度検収用）。

    Excel互換のためUTF-8 BOM付き。1セッション=1行。
    """
    rows: list[list[str]] = []
    header = [
        "session_id", "mode", "industry",
        "created_at", "updated_at", "finalized",
        "evaluation", "evaluation_correction", "evaluation_at",
        "first_user_message",
        "strategy_name", "strategy_summary",
        "output_columns", "input_table_names",
        "design_summary", "procedure_text",
        "processing_steps_count", "excel_download_url",
        "message_count", "full_transcript",
    ]
    rows.append(header)

    with _db_conn() as conn:
        if conn is None:
            raise HTTPException(503, "DB未設定")
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                "SELECT id, mode, data, created_at, updated_at "
                "FROM sessions ORDER BY created_at DESC"
            )
            db_rows = cur.fetchall()

    for r in db_rows:
        data = r.get("data") or {}
        messages = data.get("messages") or []
        first_user = next((m.get("content", "") for m in messages if m.get("role") == "user"), "")
        consult = data.get("consultation_result") or {}
        out_map = data.get("output_mapping") or {}
        out_cols = [c.get("name", "") for c in out_map.get("columns", [])] if isinstance(out_map, dict) else []
        in_tables = [t.get("table_name", "") for t in (data.get("input_tables") or []) if isinstance(t, dict)]
        design_doc = data.get("design_doc") or {}
        design_summary = design_doc.get("summary", "") if isinstance(design_doc, dict) else ""
        steps_count = len(design_doc.get("processing_steps", [])) if isinstance(design_doc, dict) else 0
        procedure_text = data.get("procedure_text", "")
        excel_url = f"/api/download/{r['id']}" if data.get("last_file") else ""
        transcript = "\n\n".join(
            f"[{m.get('role', '?')}] {m.get('content', '')}"
            for m in messages if isinstance(m, dict)
        )
        rows.append([
            r["id"],
            data.get("mode", ""),
            data.get("industry", ""),
            r["created_at"].isoformat() if r.get("created_at") else "",
            r["updated_at"].isoformat() if r.get("updated_at") else "",
            "yes" if data.get("finalized") else "no",
            data.get("evaluation", ""),
            data.get("evaluation_correction", ""),
            data.get("evaluation_at", ""),
            first_user,
            consult.get("strategy_name", "") if isinstance(consult, dict) else "",
            consult.get("strategy_summary", "") if isinstance(consult, dict) else "",
            ", ".join(out_cols),
            ", ".join(in_tables),
            design_summary,
            procedure_text,
            str(steps_count),
            excel_url,
            str(len(messages)),
            transcript,
        ])

    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_ALL)
    for row in rows:
        writer.writerow(row)
    body = "﻿" + buf.getvalue()  # UTF-8 BOM (Excelの日本語対応)
    filename = f"sessions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        content=body.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/admin/weekly_stats", dependencies=[Depends(verify_token)])
async def admin_weekly_stats():
    """週次利用集計（月〜日単位）。最大16週分を返す。"""
    from datetime import timedelta, timezone
    JST = timezone(timedelta(hours=9))
    DAYS_JA = ['月', '火', '水', '木', '金', '土', '日']

    with _db_conn() as conn:
        if conn is None:
            raise HTTPException(503, "DB未接続")
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                  date_trunc('week', created_at AT TIME ZONE 'Asia/Tokyo') AS week_start,
                  COUNT(*) FILTER (WHERE data ? 'consultation_result'
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p1_count,
                  COUNT(*) FILTER (WHERE data ? 'consultation_result' AND data->>'evaluation' = 'good'
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p1_good,
                  COUNT(*) FILTER (WHERE data ? 'consultation_result' AND data->>'evaluation' = 'bad'
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p1_bad,
                  COUNT(*) FILTER (WHERE data ? 'consultation_result' AND data->>'evaluation' IS NULL
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p1_none,
                  COUNT(*) FILTER (WHERE data->>'finalized' = 'true'
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p2_count,
                  COUNT(*) FILTER (WHERE data->>'finalized' = 'true' AND data->>'evaluation' = 'good'
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p2_good,
                  COUNT(*) FILTER (WHERE data->>'finalized' = 'true' AND data->>'evaluation' = 'bad'
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p2_bad,
                  COUNT(*) FILTER (WHERE data->>'finalized' = 'true' AND data->>'evaluation' IS NULL
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p2_none,
                  COUNT(*) FILTER (WHERE data ? 'design_doc'
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p3_count,
                  COUNT(*) FILTER (WHERE data ? 'design_doc' AND data->>'evaluation' = 'good'
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p3_good,
                  COUNT(*) FILTER (WHERE data ? 'design_doc' AND data->>'evaluation' = 'bad'
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p3_bad,
                  COUNT(*) FILTER (WHERE data ? 'design_doc' AND data->>'evaluation' IS NULL
                                     AND (data->>'is_test' IS NULL OR data->>'is_test' != 'true'))             AS p3_none
                FROM sessions
                GROUP BY week_start
                ORDER BY week_start DESC
                LIMIT 16
            """)
            rows = cur.fetchall()

    # 累計計算（古い順に積算）
    rows_asc = list(reversed(rows))
    cum = [0, 0, 0]
    result = []
    for row in rows_asc:
        ws = row[0]
        if ws is None:
            continue
        ws_jst = ws.replace(tzinfo=timezone.utc).astimezone(JST)
        we_jst = ws_jst + timedelta(days=6)
        label = (f"{ws_jst.month}/{ws_jst.day}({DAYS_JA[ws_jst.weekday()]})"
                 f"～{we_jst.month}/{we_jst.day}({DAYS_JA[we_jst.weekday()]})")
        p1, p2, p3 = int(row[1]), int(row[5]), int(row[9])
        result.append({
            "week": label,
            "p1": {"count": p1, "good": int(row[2]),  "bad": int(row[3]),  "none": int(row[4])},
            "p2": {"count": p2, "good": int(row[6]),  "bad": int(row[7]),  "none": int(row[8])},
            "p3": {"count": p3, "good": int(row[10]), "bad": int(row[11]), "none": int(row[12])},
        })

    return list(reversed(result))


@app.get("/api/admin/weekly_stats/csv", dependencies=[Depends(verify_token)])
async def admin_weekly_stats_csv():
    """週次集計をCSVでダウンロード"""
    stats = await admin_weekly_stats()
    header = (
        "週,"
        "a.要件定義_利用回数,a.要件定義_good,a.要件定義_bad,a.要件定義_未リアクション,"
        "b.テーブル定義書_利用回数,b.テーブル定義書_good,b.テーブル定義書_bad,b.テーブル定義書_未リアクション,"
        "c.構築手順書_利用回数,c.構築手順書_good,c.構築手順書_bad,c.構築手順書_未リアクション"
    )
    lines = [header]
    for r in stats:
        lines.append(
            f"{r['week']},"
            f"{r['p1']['count']},{r['p1']['good']},{r['p1']['bad']},{r['p1']['none']},"
            f"{r['p2']['count']},{r['p2']['good']},{r['p2']['bad']},{r['p2']['none']},"
            f"{r['p3']['count']},{r['p3']['good']},{r['p3']['bad']},{r['p3']['none']}"
        )
    body = "﻿" + "\n".join(lines)  # BOM付きUTF-8（Excel対応）
    return Response(
        content=body.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=\"weekly_stats.csv\""},
    )


@app.get("/api/admin/sessions", dependencies=[Depends(verify_token)])
async def admin_list_sessions(limit: int = 100):
    """이력 조회: design_doc이 있는 세션 목록을 최신순으로 반환"""
    with _db_conn() as conn:
        if conn is None:
            raise HTTPException(503, "DB未接続")
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT
                    id,
                    created_at,
                    mode,
                    data->'consultation_result'->>'strategy_name' AS strategy_name,
                    data->'design_doc'->>'summary'               AS summary,
                    data ? 'design_doc'                          AS has_output,
                    (data->>'is_test' = 'true')                  AS is_test
                FROM sessions
                WHERE data ? 'design_doc'
                ORDER BY created_at DESC
                LIMIT %s
            """, (limit,))
            rows = cur.fetchall()
    return [dict(r) for r in rows]


@app.post("/api/admin/sessions/{session_id}/toggle_test", dependencies=[Depends(verify_token)])
async def admin_toggle_test(session_id: str):
    """セッションのテスト除外フラグをトグルする"""
    with _db_conn() as conn:
        if conn is None:
            raise HTTPException(503, "DB未接続")
        with conn.cursor() as cur:
            cur.execute("SELECT data FROM sessions WHERE id = %s", (session_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "セッションが見つかりません")
            data = row[0]
            new_val = not (data.get("is_test") is True)
            data["is_test"] = new_val
            cur.execute(
                "UPDATE sessions SET data = %s, updated_at = NOW() WHERE id = %s",
                (Json(data), session_id),
            )
        conn.commit()
    # メモリ上のセッションにも反映
    mem = sessions.get(session_id)
    if mem:
        mem["is_test"] = new_val
        sessions.save(session_id, mem)
    return {"session_id": session_id, "is_test": new_val}


@app.get("/api/admin/sessions/{session_id}", dependencies=[Depends(verify_token)])
async def admin_get_session(session_id: str):
    """이력 상세: 대화내용 + design_doc 전체 반환 (DB優先、フォールバックでメモリ)"""
    session = None
    with _db_conn() as conn:
        if conn is not None:
            with conn.cursor() as cur:
                cur.execute("SELECT data FROM sessions WHERE id = %s", (session_id,))
                row = cur.fetchone()
                if row:
                    session = row[0]
    if session is None:
        session = sessions.get(session_id)
    if not session:
        raise HTTPException(404, "セッションが見つかりません")
    cr = session.get("consultation_result") or {}
    return {
        "session_id": session_id,
        "strategy_name": cr.get("strategy_name", ""),
        "summary": (session.get("design_doc") or {}).get("summary", ""),
        "messages": session.get("messages", []),
        "messages_step2": session.get("messages_step2", []),
        "consultation_messages": session.get("consultation_messages", []),
        "consultation_result": cr,
        "design_doc": session.get("design_doc"),
        "input_tables": session.get("input_tables", []),
        "output_mapping": session.get("output_mapping", {}),
        "has_download": bool(session.get("design_doc")),
    }


@app.get("/admin", response_class=HTMLResponse)
async def admin_page():
    html_path = FRONTEND_PATH / "admin.html"
    html = html_path.read_text(encoding="utf-8")
    token_script = f'<script>window.__APP_AUTH_TOKEN__="{AUTH_TOKEN}";</script>'
    html = html.replace("</head>", f"{token_script}</head>")
    return html


@app.get("/healthz")
async def healthz():
    """ALB/ECSヘルスチェック用。認証・DB・外部API非依存で即200を返す。"""
    return {"status": "ok"}


@app.get("/", response_class=HTMLResponse)
async def index():
    html_path = FRONTEND_PATH / "index.html"
    html = html_path.read_text(encoding="utf-8")
    # Bearer Token をフロントに注入（全fetchリクエストで自動送信させる）
    token_script = f'<script>window.__APP_AUTH_TOKEN__="{AUTH_TOKEN}";</script>'
    html = html.replace("</head>", f"{token_script}</head>")
    return html


@app.get("/bdash_hakase.png")
async def avatar_image():
    return FileResponse(FRONTEND_PATH / "bdash_hakase.png", media_type="image/png")

@app.get("/favicon.png")
async def favicon():
    return FileResponse(FRONTEND_PATH / "favicon.png", media_type="image/png")
