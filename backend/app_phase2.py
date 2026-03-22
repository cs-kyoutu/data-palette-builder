"""
Phase2: 手順書ジェネレーター
Phase1で作成した設計書JSONを入力に、完全な手順書を一発生成する。
質問は一切せず、設計書の内容に基づいて詳細な手順書をExcel形式で出力。
"""
import json
import os
import uuid
from datetime import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
except ImportError:
    pass

import anthropic
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel

app = FastAPI(title="手順書ジェネレーター（Phase2）")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).parent.parent
OUTPUT_DIR = Path(__file__).parent / "output"
UPLOAD_DIR = Path(__file__).parent / "uploads"
OUTPUT_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)

sessions: dict[str, dict] = {}
client = anthropic.Anthropic()


class GenerateRequest(BaseModel):
    session_id: str | None = None
    design_document: dict

class GenerateResponse(BaseModel):
    session_id: str
    status: str  # "done" | "error"
    message: str
    download_url: str | None = None


# --- Phase2用システムプロンプト ---
SYSTEM_PROMPT_PHASE2 = """あなたはb→dashのデータパレット構築手順書を生成する専門AIです。

## あなたの役割
提供された「設計書」に基づいて、**完全かつ詳細な手順書**を生成してください。
設計書にはすべての意思決定が含まれているため、**質問は一切不要**です。
**省略せず、すべてのステップを完全に記述**してください。

## 設計書の内容
{design_document}

## b→dashデータパレットの操作仕様（正式名称・制約含む）

### ===== 統合（2種類） =====

#### 横統合（テーブル結合）
- **定義**: 2つのデータファイルで共通するカラムをキーに、1つのデータファイルにまとめる
- **操作パス**: データパレット → データを確認する → 統合する → カスタマイズ → ファイル選択 → 横統合
- **制約**: **3つ以上のデータファイルを同時に横統合できない**（2ファイルずつ）
- **統合方法（4パターン）**:
  1. 「全てのデータを統合する」= FULL OUTER JOIN相当
  2. 「共通のデータのみを統合する」= INNER JOIN
  3. 「先に選択したデータに対して統合する」= LEFT JOIN
  4. 「後に選択したデータに対して統合する」= RIGHT JOIN
- **キーの注意**: NULL同士は統合されない。空文字キーは多対多で膨大レコード生成リスク
- **設定項目**: 統合キー選択、残すカラム選択、統合データ名、キー重複時の処理、更新設定

#### 縦統合（レコード結合）
- **定義**: 異なるデータファイル同士で共通するカラムのレコードを統合する（UNION相当）
- **操作パス**: データパレット → データを確認する → 統合する → カスタマイズ → 縦統合
- **制約**: **最大4ファイル**同時統合可能
- **設定項目**: 統合カラム指定、重複データ除外する/しない、更新設定

### ===== 加工（21種類） =====

すべての加工の**共通操作パス**: データパレット → データを確認する → データファイル選択 → 加工する → 編集方法選択 → [加工種別]選択
すべての加工の**共通制約**: 上書き保存は「過去1回以上クレンジング済みファイル」のみ。未加工・統合ファイルは「新規データファイルとして保存」のみ。

#### 1. 連結
- カラムを2つ選択して1つに結合。テキスト挿入で区切り文字（"/"、スペース等）挿入可能
- 元カラム残す/残さない選択可

#### 2. テキスト挿入
- カラムの文頭または文末に任意テキストを挿入

#### 3. 分割
- 区切り文字を指定してカラムを複数カラムに分割
- 設定: 区切り文字、開始位置、分割数
- 注意: 小数型カラムを「.」で分割すると小数点なし値は整数と空文字に分割される

#### 4. 四則演算
- +/-/×/÷ の計算。端数処理（四捨五入/切り捨て/切り上げ、小数第1～4位）
- NULL含有時の結果はNULL（テキスト型）

#### 5. 時刻演算
- 日付/日時データの加減算。右辺: 本日の日付/カスタム/特定の日付
- NULL含有時の結果はNull
- 左辺日時型: 日付型/日時型との引き算のみ。整数型との足し引き可

#### 6. IF文
- 条件分岐で値を変換。複数条件設定可（IF-ELSE IF構造）
- 操作: カラム選択 → 分岐条件と値を入力 → 追加 → 繰り返し

#### 7. 追加（カラム追加）
- 任意の位置に新カラム追加。左に追加/右に追加
- データ型: テキスト型、数値型、日付型、日時型
- 「加工処理実行日カラム」チェックボックスあり（日付/日時型選択時）

#### 8. 複製
- 既存カラムのコピーを作成

#### 9. 削除
- カラム削除（複数選択で一括削除可能）

#### 10. ランキング
- 順位付け。パターン: 「1列を選択して順位付け」/「グループ内で順位付け」
- グループ化カラム選択、昇順/降順、同率順位の扱い選択

#### 11. 集約
- グループ化して集計。グループ化カラム（複数選択可）、集計カラム（複数選択可）
- 集計方法: カウント/ユニークカウント/合計/平均/最大値/最小値/最新日時/最古日時/行結合
- データ型制限: 合計・平均・最大値・最小値は数値型のみ。最新/最古日時は日付・日時型のみ

#### 12. 置換
- カラムの値を条件指定で変換。複数置換ルール設定可
- 名前をつけて保存（元カラム残存）/ 上書き保存（元カラム削除）

#### 13. 型変換
- データ型を変換。変換元の形式を最大5個選択可（日付/日時変換時は必須）
- 日付型: 33フォーマット対応。空文字→NULL自動変換

#### 14. 抽出
- 3パターン:
  - **先頭から抽出**: X文字目までを抽出（例: 先頭5文字で"2026-"）
  - **中間を抽出**: X〜Y文字目までを抽出
  - **末尾から抽出**: 最後からX文字目までを抽出（例: 末尾5文字で"03-15"）

#### 15. 除外
- カラムから特定の文字列を削除

#### 16. 書式変換
- 全角/半角、大文字/小文字の変換

#### 17. 0埋め
- 先頭に0挿入して桁数統一

#### 18. 絞込み
- 条件指定でレコードをフィルタリング。AND/OR条件組み合わせ可能

#### 19. 名寄せ（重複排除）
- キーカラムの重複レコードを優先順位に応じて1件に絞り込み
- 設定: キーカラム、優先順位カラム（例: 最新日付を優先）

#### 20. 参照
- グループ内でソート後に特定の値を取得
- 6パターン: 最初の値/最後の値/特定順番の値/前の行の値/後ろの行の値/累計の値
- 設定: グループ化カラム、ソートカラム・順序、値取得カラム

#### 21. 並び替え
- カラムの表示順序をドラッグ&ドロップで変更

### ===== 加工テンプレート（4種類） =====

加工テンプレートは、複数のクレンジング操作を組み合わせた事前定義済み処理。
**共通操作パス**: データパレット → データを確認する → データファイル選択 → 加工する → テンプレート → [テンプレート名]を選択 → 適用

#### T1. 横持ちを縦に並べて変換（横→縦変換 / UNPIVOT相当）
- **定義**: 横方向に展開された複数カラムを縦方向に統合する
- **操作**: テンプレート → 「顧客ごとに横持ちのデータを、縦に並べて変換」を選択
- **設定項目**:
  - 縦持ちにしたいカラムを複数選択（最大**50個**）
  - 残すカラムを選択（最大**50個**）
  - 新カラム名を編集
- **例**: 購入商品(1), 購入商品(2), 購入商品(3) → 1カラム「縦持ちカラム」に縦展開
- **操作種別の記載**: 「テンプレート\n横持ちを縦持ちに変換」

#### T2. 縦持ちを横に並べて変換（縦→横変換 / PIVOT相当）
- **定義**: 縦方向の複数レコードを横方向に展開する
- **操作**: テンプレート → 「顧客ごとに縦持ちのデータを、横に並べて変換」を選択
- **設定項目**:
  - 集約キーカラム選択（最大**20個**）
  - 横並びに変換するカラム選択（最大**20個**）
  - 並び順カラム選択（最大**5個**）＋昇順/降順
  - 何番目まで並べるか（**1〜15個**）
  - カラム名の編集
- **例**: 顧客IDをキーに、商品ID/商品名/価格を横展開、購入日降順で上位5件
- **操作種別の記載**: 「テンプレート\n縦持ちを横持ちに変換」

#### T3. 都道府県を地域に変換
- 住所データから都道府県を抽出し、地域カテゴリ（関東、関西等）に分類
- 対象カラム選択 → 新カラム名設定

#### T4. 生年月日から年齢を算出
- 生年月日カラムから現在の年齢を自動計算して新カラム追加
- 対象カラム選択 → 新カラム名設定

#### T5. 金額をカンマ区切りに変換
- 「12500」→「12,500」のように金額にカンマ挿入
- **制約**: 整数型カラムのみ適用可能

## 手順書生成のルール

### 絶対に守ること
1. **省略禁止**: 「同様に〜」「以下同様」のような省略は禁止。すべてのステップを個別に記述
2. **b→dashのUI操作名を正確に使う**: 「横統合」「絞込み」「集約」「名寄せ」「参照」「分割」「連結」「時刻演算」等の正式名称
3. **操作パスを明記**: 「データパレット → データを確認する → [ファイル名] → 加工する → 編集方法を選択 → [加工種別]」
4. **中間ファイル名を明記**: 連番＋内容で命名（例: 「01_アクセスログ×顧客_横統合」）
5. **更新設定の方針**: 最終ステップのみ「更新設定をする」、途中は「更新設定をしない」
6. **横統合の統合方法はb→dash名称で**: 「共通のデータのみを統合する」「先に選択したデータに対して統合する」
7. **残すカラムを明示**: 横統合時は残すカラムを具体的にリスト
8. **同一ファイル内の連続加工**: Step番号は空にして保存ファイル名も空、最後の加工時のみ保存
9. **操作種別に「テンプレート」と記載可**: 「テンプレート\n横持ちを縦持ちに変換」「テンプレート\n縦持ちを横持ちに変換」
10. **ステップ数は最小限に**: 冗長な繰り返しは避ける。参照は1回でキーにするカラムを複合指定する

### 各ステップの記載項目（8カラム）
Step | 操作種別(b→dash名) | UI操作パス | 操作内容・設定値 | 保存ファイル名 | 結果の状態 | 確認ポイント | 備考

## 出力形式
必ず以下のJSON形式で出力してください（```json で囲む）：

```json
{{
  "action": "generate",
  "title": "手順書タイトル",
  "sections": [
    {{
      "sheet_name": "概要",
      "title": "データパレット構築 概要",
      "columns": ["項目", "内容"],
      "rows": [
        ["目的", "○○テーブルの構築"],
        ["インプットテーブル", "テーブルA, テーブルB, ..."],
        ["アウトプットカラム数", "XX カラム"],
        ["処理ステップ数", "XX ステップ"],
        ["横統合回数", "X 回"],
        ["加工回数", "X 回"],
        ["データフロー概要", "処理の全体像を記述"]
      ]
    }},
    {{
      "sheet_name": "データ準備",
      "title": "Step1: データ準備",
      "columns": ["No", "データファイル名", "用途", "使用カラム", "備考"],
      "rows": [...]
    }},
    {{
      "sheet_name": "結合・加工",
      "title": "Step2: 結合・加工手順",
      "columns": ["Step", "操作種別(b→dash名)", "使用データ", "UI操作パス", "操作内容・設定値", "保存ファイル名", "結果の状態", "確認ポイント", "備考"],
      "rows": [
        ["1", "横統合", "webアクセスログデータ, 顧客データ", "データパレット → データを確認する → 統合する → カスタマイズ → webアクセスログテーブルと顧客テーブルを選択 → 横統合", "統合方法: 共通のデータのみを統合する\\n統合キー: リレーション項目_1 = 顧客ID\\n残すカラム: PV/Click日時, リレーション項目_1, リレーション項目_10, メールアドレス(メイン), メルマガ配信許可フラグ, 姓, 名", "01_アクセスログ×顧客_横統合", "顧客マスタと一致するアクセスログのみが残る", "結合前後のレコード数、NULLレコード数を確認", "更新設定: しない"],
        ["2", "絞込み", "01_アクセスログ×顧客_横統合", "データパレット → データを確認する → 01_アクセスログ×顧客_横統合 → 加工する → 編集方法を選択 → 絞込み", "条件1: PV/Click日時 次の期間にある 相対期間 3月前 0日前\\nAND\\n条件2: リレーション項目_10 空文字ではない", "02_直近3ヶ月カート投入_絞込み", "直近3ヶ月のカート投入データのみが残る", "期間内レコード数を確認", "更新設定: しない"],
        ["3", "分割", "02_直近3ヶ月カート投入_絞込み", "データパレット → データを確認する → 02_直近3ヶ月カート投入_絞込み → 加工する → 編集方法を選択 → 分割", "対象カラム: リレーション項目_10\\n区切り文字: ,\\n左から\\n分割数: 10\\n新カラム名: 商品ID_1, 商品ID_2, ..., 商品ID_10", "03_商品ID分割", "カンマ区切りの商品IDが個別カラムに分割される", "分割後の各カラムの値分布を確認", "更新設定: しない"],
        ["", "参照", "03_商品ID分割", "（同ファイル連続加工）→ 参照", "キーにするカラム: リレーション項目_1, 縦持ちカラム\\nソートカラム: PV/Click日時\\nソート順: 最も新しい日時\\n値取得カラム: PV/Click日時\\n新カラム名: 最終カート投入日時", "", "", "", "同一ファイルで連続加工"],
        ["4", "テンプレート\\n横持ちを縦持ちに変換", "03_商品ID分割", "データパレット → データを確認する → 03_商品ID分割 → 加工する → テンプレート → 横持ちを縦持ちに変換", "縦持ちにするカラム: 商品ID_1, 商品ID_2, ..., 商品ID_10\\n残すカラム: PV/Click日時, リレーション項目_1, メールアドレス(メイン), メルマガ配信許可フラグ, 姓, 名", "04_商品ID縦持ち変換", "各商品IDが個別行に展開される", "展開後のレコード数を確認", "更新設定: しない"],
        ["5", "横統合", "04_商品ID縦持ち変換, 商品データ", "データパレット → データを確認する → 統合する → カスタマイズ → 04_商品ID縦持ち変換と商品テーブルを選択 → 横統合", "統合方法: 先に選択したデータに対して統合する\\n統合キー: 縦持ちカラム = 商品ID\\n残すカラム: 全カラム", "05_商品情報結合", "商品情報が結合される", "商品情報の結合率、NULL件数を確認", "更新設定: しない"],
        ["6", "集約", "05_商品情報結合", "データパレット → データを確認する → 05_商品情報結合 → 加工する → 編集方法を選択 → 集約", "グループ化カラム: リレーション項目_1\\n集計: PV/Click日時 → 最新日時 → 最終カート投入日\\n残すカラム: メールアドレス(メイン), メルマガ配信許可フラグ, 姓, 名", "", "", "", "同一ファイルで連続加工"],
        ["", "連結", "05_商品情報結合", "（同ファイル連続加工）→ 連結", "連結カラム1: 姓\\n連結カラム2: 名\\n区切り文字: スペース\\n新カラム名: 名前\\n元カラム: 残さない", "", "", "", ""],
        ["", "時刻演算", "05_商品情報結合", "（同ファイル連続加工）→ 時刻演算", "左辺カラム: 本日の日付\\n演算子: -\\n右辺カラム: 最終カート投入日\\n単位: 時間\\n新カラム名: 最終カート投入日からの経過時間", "06_顧客集約加工済み", "名前結合＋経過時間計算が完了", "経過時間の妥当性を確認", "更新設定: しない"],
        ["7", "テンプレート\\n縦持ちを横持ちに変換", "05_商品情報結合", "データパレット → データを確認する → 05_商品情報結合 → 加工する → テンプレート → 縦持ちを横持ちに変換", "キーにするカラム: リレーション項目_1, 最終カート投入日, 最終カート投入日からの経過時間, メールアドレス(メイン), メルマガ配信許可フラグ, 名前\\n横持ちカラム: 商品ID, 商品名, 商品詳細ページURL, 商品価格, 商品画像URL\\nソート順: 降順\\nソートするカラム: PV/Click日時\\n何番目まで並べるか: 上位5位まで", "07_カート放棄施策配信リスト", "顧客ごとに上位5商品が横展開される", "ランキングの付与状況を確認", "更新設定: する"]
      ]
    }},
    {{
      "sheet_name": "最終確認",
      "title": "Step3: 最終確認チェックリスト",
      "columns": ["No", "確認項目", "アウトプットカラム", "ソース", "加工方法", "期待値", "確認結果", "備考"],
      "rows": [...]
    }},
    {{
      "sheet_name": "検証観点",
      "title": "Step4: 検証観点",
      "columns": ["No", "検証カテゴリ", "検証項目", "検証方法", "期待結果", "実際結果", "OK/NG", "対処方法"],
      "rows": [...]
    }}
  ]
}}
```

## 重要な設計パターン

### パターン1: 参照は1回でまとめる
商品ID_1〜10のようにカラムが複数あっても、参照は**1回**で実行する。キーにするカラムに複合キー（例: リレーション項目_1, 縦持ちカラム）を指定する。**10回繰り返さない**。

### パターン2: 同一ファイルの連続加工（まとめられるものはまとめる）
集約→連結→時刻演算のように同一ファイルに対する加工が続く場合は、**1回の加工操作としてまとめる**。
毎回ファイル保存せず、最後にまとめて保存する。
- 最初の加工にStep番号を記載、最後の加工に保存ファイル名を記載
- 途中の加工はStep番号を空、保存ファイル名も空にする
- UI操作パスは「（同ファイル連続加工）→ 連結」のように簡潔に

### パターン3: ステップ数は最小限
上記参考例のように**7〜9ステップ程度**で完結させる。不要な中間ステップは作らない。

### パターン4: 使用データ列の書き方
- 横統合の場合: 「webアクセスログデータ, 顧客データ」のように元テーブル名をカンマ区切り
- 加工の場合: 加工元の中間ファイル名（例: 「01_アクセスログ×顧客_横統合」）
- 連続加工の場合: 同じファイル名を繰り返し記載

### パターン5: カラム値同士の比較（絞込みではできない）
b→dashの絞込みには「カラムAの値がカラムBの値と一致する」という条件がない。
カラム同士を比較してフィルタしたい場合は以下の2段階で実現する：
1. **IF文**: 対象カラム選択 → 条件「次のカラムの値と完全一致」→ カラム選択 → 一致時"1"、不一致時""（空文字）→ 新カラム名「判定フラグ」
2. **絞込み**: 判定フラグが空文字のレコードを残す（または"1"のレコードを残す）
**絞込みで直接カラム同士を比較する設定は絶対に使わないこと。**

検証観点シートには以下を含めてください：
- 各横統合ステップ前後のレコード件数比較
- 必須カラムごとのNULL件数確認
- キーカラムの一意性チェック
- 横統合キーの一致率（結合できなかったレコード数）
- 値域チェック（数値範囲、日付妥当性）
- サンプル3-5件の目視確認

## 設計書v2.0のprocessing_stepsがある場合
設計書に`processing_steps`が含まれている場合、**それをほぼそのまま「結合・加工」シートの各行に変換**してください：
- step → Step列（nullの場合は空文字）
- operation → 操作種別列
- **使用データ列**: 横統合の場合は元テーブル名をカンマ区切り（例:「顧客データ, 受注データ」）、加工の場合は加工元ファイル名
- ui_path → UI操作パス列
- settings → 操作内容・設定値列（settingsの各項目を改行区切りテキストに変換）
- save_as → 保存ファイル名列
- result → 結果の状態列
- check → 確認ポイント列
- note → 備考列

processing_stepsがある場合はAIが処理ロジックを再設計する必要はありません。設計書を忠実にExcel形式に変換するだけです。

## 重要
- **質問は絶対にしないでください。** 設計書にすべての情報があります。
- 不明点があっても推論して最善の手順を生成してください。
- 手順書は**b→dashを操作する実務担当者がこの手順書だけで作業完了できるレベル**の具体性で記述してください。
- SQLは使用しないでください。すべてb→dashのGUI操作（統合・加工メニュー）で実現してください。
"""


def format_design_document(doc: dict) -> str:
    """設計書JSONを読みやすいテキストに変換"""
    lines = []

    # 概要
    if "summary" in doc:
        lines.append(f"### 概要\n{doc['summary']}\n")

    # インプットテーブル
    if "input_tables" in doc:
        lines.append("### インプットテーブル")
        for table in doc["input_tables"]:
            lines.append(f"\n#### {table.get('table_name', 'unknown')}テーブル")
            lines.append("| カラム名 | 型 | 説明 |")
            lines.append("|---------|-----|------|")
            for col in table.get("columns", []):
                lines.append(f"| {col.get('name', '')} | {col.get('type', '')} | {col.get('description', '')} |")
        lines.append("")

    # アウトプットマッピング
    if "output_mapping" in doc:
        lines.append("### アウトプットマッピング")
        lines.append("| カラム名 | 定義 | ソースカラム | ソーステーブル | 導出方法 |")
        lines.append("|---------|------|------------|-------------|---------|")
        for col in doc["output_mapping"].get("columns", []):
            lines.append(
                f"| {col.get('name', '')} | {col.get('definition', '')} | "
                f"{col.get('source_column', '') or '加工'} | {col.get('source_table', '')} | "
                f"{col.get('derivation', '')} |"
            )
        lines.append("")

    # v2.0: processing_steps（b→dash操作レベルの処理ステップ）
    if "processing_steps" in doc:
        lines.append("### 処理ステップ（b→dash操作レベル）")
        lines.append("以下のステップをそのまま「結合・加工」シートに変換してください：\n")
        for ps in doc["processing_steps"]:
            step_num = ps.get("step", "")
            step_label = f"Step {step_num}" if step_num else "  (連続加工)"
            lines.append(f"#### {step_label}: {ps.get('operation', '')}")
            lines.append(f"- UI操作パス: {ps.get('ui_path', '')}")
            settings = ps.get("settings", {})
            if settings:
                lines.append("- 設定値:")
                for k, v in settings.items():
                    if isinstance(v, list):
                        lines.append(f"  - {k}: {', '.join(str(x) for x in v)}")
                    elif isinstance(v, dict):
                        lines.append(f"  - {k}: {json.dumps(v, ensure_ascii=False)}")
                    else:
                        lines.append(f"  - {k}: {v}")
            if ps.get("save_as"):
                lines.append(f"- 保存ファイル名: {ps['save_as']}")
            if ps.get("result"):
                lines.append(f"- 結果: {ps['result']}")
            if ps.get("check"):
                lines.append(f"- 確認: {ps['check']}")
            if ps.get("note"):
                lines.append(f"- 備考: {ps['note']}")
            lines.append("")

    # v1.0互換: 意思決定
    decisions = doc.get("decisions", {})

    if "joins" in decisions:
        lines.append("### 結合戦略")
        for j in decisions["joins"]:
            lines.append(
                f"- Step {j.get('step_order', '?')}: {j.get('left_table', '')} と {j.get('right_table', '')} を "
                f"{j.get('join_type', 'JOIN')}（キー: {j.get('join_key_left', '')} = {j.get('join_key_right', '')}）"
                f" → {j.get('reason', '')}"
            )
        lines.append("")

    if "transformations" in decisions:
        lines.append("### 加工ロジック")
        for t in decisions["transformations"]:
            lines.append(
                f"- Step {t.get('step_order', '?')}: [{t.get('type', '')}] "
                f"{t.get('detail', '')}（出力: {t.get('output_column', '')}）"
            )
        lines.append("")

    if "filters" in decisions:
        lines.append("### フィルタ条件")
        for f in decisions["filters"]:
            lines.append(f"- {f.get('column', '')} {f.get('operator', '')} {f.get('value', '')} → {f.get('reason', '')}")
        lines.append("")

    if "processing_order" in decisions:
        lines.append("### 処理順序")
        for step in decisions["processing_order"]:
            lines.append(f"- {step}")
        lines.append("")

    if "null_handling" in decisions:
        lines.append("### NULL処理方針")
        for n in decisions["null_handling"]:
            lines.append(f"- {n.get('column', '')}: {n.get('strategy', '')} {n.get('default_value', '') or ''}")
        lines.append("")

    if "special_notes" in decisions:
        lines.append("### 特記事項")
        for note in decisions["special_notes"]:
            lines.append(f"- {note}")
        lines.append("")

    # Q&A履歴
    if "qa_history" in doc:
        lines.append("### Q&A履歴（ヒアリング結果）")
        for qa in doc["qa_history"]:
            lines.append(f"- Q: {qa.get('question', '')}")
            lines.append(f"  A: {qa.get('answer', '')}")
            lines.append(f"  影響: {qa.get('impact', '')}")
        lines.append("")

    return "\n".join(lines)


def build_spreadsheet(generation_data: dict) -> tuple[str, str]:
    """生成データから1シート形式のExcelファイルを作成（参考スプレッドシート準拠）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "手順書"

    # --- スタイル定義 ---
    GRAY_BG = "EFEFEF"
    font9 = Font(name="Yu Gothic UI", size=9)
    font9_bold = Font(name="Yu Gothic UI", size=9, bold=True)
    font10 = Font(name="Yu Gothic UI", size=10)
    section_fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    STEP_MARKS = ["①", "②", "③", "④", "⑤", "⑥", "⑦", "⑧", "⑨", "⑩",
                  "⑪", "⑫", "⑬", "⑭", "⑮", "⑯", "⑰", "⑱", "⑲", "⑳"]

    # --- 列幅設定（均一グリッド） ---
    ws.column_dimensions["A"].width = 4
    for c in "BCDEFGHIJKLMNOPQRSTU":
        ws.column_dimensions[c].width = 13

    # --- セクション帯を書く関数 ---
    def write_section_header(row, title_text):
        for col in range(1, 21):  # A-T
            cell = ws.cell(row=row, column=col)
            cell.fill = section_fill
        ws.cell(row=row, column=2, value=title_text).font = font9_bold

    # --- セクション探索 ---
    sections = generation_data.get("sections", [])
    overview_sec = next((s for s in sections if s.get("sheet_name") == "概要"), None)
    prep_sec = next((s for s in sections if "準備" in s.get("sheet_name", "")), None)
    proc_sec = next((s for s in sections if "加工" in s.get("sheet_name", "") or "結合" in s.get("sheet_name", "")), None)
    check_sec = next((s for s in sections if "確認" in s.get("sheet_name", "")), None)
    verify_sec = next((s for s in sections if "検証" in s.get("sheet_name", "")), None)

    # --- 3x3ボックスを描画する関数 ---
    box_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    box_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

    def draw_box(row, col, label, id_label="ID"):
        """3x3マージセルのボックスを描画。上にIDラベル付き"""
        # ID ラベル（ボックス上部、マージ）
        ws.cell(row=row, column=col, value=id_label).font = font9
        if col + 1 <= 20:
            ws.merge_cells(start_row=row, start_column=col + 1,
                          end_row=row, end_column=col + 2)
        # ボックス本体（3行×3列マージ）
        box_cell = ws.cell(row=row + 1, column=col, value=label)
        box_cell.font = font10
        box_cell.alignment = box_align
        ws.merge_cells(start_row=row + 1, start_column=col,
                      end_row=row + 3, end_column=col + 2)
        # ボーダー
        for r in range(row + 1, row + 4):
            for c in range(col, col + 3):
                ws.cell(row=r, column=c).border = box_border

    cur_row = 1

    # ========== ■ フロー セクション ==========
    write_section_header(cur_row, "■ フロー")
    cur_row += 2  # 空行

    # フロー図：テーブルとステップの関係を可視化
    if prep_sec and prep_sec.get("rows") and proc_sec and proc_sec.get("rows"):
        tables = prep_sec["rows"]
        proc_rows = proc_sec["rows"]

        # メインステップ抽出
        main_steps = [r for r in proc_rows if r[0] and str(r[0]).strip()]

        # テーブル名リスト
        tbl_names = [t[1] if len(t) > 1 else f"テーブル{i}" for i, t in enumerate(tables)]

        # どのステップでどのテーブルが初登場するか判定
        tbl_first_step = {}
        for t_name in tbl_names:
            for s_idx, step in enumerate(main_steps):
                combined = " ".join(str(step[c]) for c in range(len(step)) if step[c])
                if t_name in combined:
                    tbl_first_step[t_name] = s_idx
                    break
            if t_name not in tbl_first_step:
                tbl_first_step[t_name] = 0

        # Step1で使われるテーブル → 上段の先頭に配置
        upper_tables = [t for t in tbl_names if tbl_first_step.get(t, 0) == 0]
        lower_tables = [(t, tbl_first_step[t]) for t in tbl_names if tbl_first_step.get(t, 0) > 0]

        flow_id_row = cur_row
        base_col = 3  # C列

        # 上段: 最初のテーブル → ①→②→③...
        if upper_tables:
            draw_box(flow_id_row, base_col, upper_tables[0])

        # ステップボックス（5列間隔で横に配置）
        step_col_map = {}
        for s_idx in range(len(main_steps)):
            s_col = base_col + 5 + s_idx * 5
            if s_col + 2 > 21:
                break
            mark = STEP_MARKS[s_idx] if s_idx < len(STEP_MARKS) else f"({s_idx+1})"
            draw_box(flow_id_row, s_col, mark)
            step_col_map[s_idx] = s_col

            # 横矢印（→）
            arrow_col = s_col - 1
            if arrow_col > base_col + 2:
                ws.cell(row=flow_id_row + 2, column=arrow_col, value="→").font = font10

        # 最初のテーブル→①の矢印
        if step_col_map:
            first_step_col = step_col_map[0]
            ws.cell(row=flow_id_row + 2, column=first_step_col - 1, value="→").font = font10

        # 下段: 後から結合されるテーブルを、対応するステップの下に配置
        if lower_tables:
            lower_row = flow_id_row + 5
            for t_name, s_idx in lower_tables:
                t_col = step_col_map.get(s_idx, base_col)
                draw_box(lower_row, t_col, t_name)
                # 上矢印（↑）：下段テーブル → 上段ステップ
                ws.cell(row=lower_row - 1, column=t_col + 1, value="↑").font = font10
            cur_row = lower_row + 5
        else:
            cur_row = flow_id_row + 5

    cur_row += 1  # 空行

    # ========== ■ 手順書 セクション ==========
    write_section_header(cur_row, "■ 手順書")
    cur_row += 1

    # ヘッダー行（H列: 対象データ、I列: 加工内容）
    ws.cell(row=cur_row, column=8, value="対象データ").font = font9_bold
    ws.cell(row=cur_row, column=9, value="加工内容").font = font9_bold
    cur_row += 1

    if proc_sec and proc_sec.get("rows"):
        proc_rows = proc_sec["rows"]

        # columns から列インデックスを特定
        columns = proc_sec.get("columns", [])
        def col_idx(name):
            for i, c in enumerate(columns):
                if name in c:
                    return i
            return -1

        idx_step = col_idx("Step")
        idx_op = col_idx("操作種別")
        idx_used = col_idx("使用データ")
        idx_settings = col_idx("操作内容")
        if idx_settings < 0:
            idx_settings = col_idx("加工内容")
        idx_save = col_idx("保存ファイル")

        step_counter = 0  # ①②③ のカウンタ
        sub_counter = 0   # サブ番号

        for p_idx, p_row in enumerate(proc_rows):
            def get(idx):
                if idx >= 0 and idx < len(p_row):
                    return str(p_row[idx]) if p_row[idx] else ""
                return ""

            step_val = get(idx_step)
            op_type = get(idx_op)
            used_data = get(idx_used)
            settings = get(idx_settings)

            # 新ステップかサブステップか判定
            is_new_step = step_val and step_val.strip()

            if is_new_step:
                cur_row += 1  # ステップ間に空行
                step_counter += 1
                sub_counter = 1
                mark = STEP_MARKS[step_counter - 1] if step_counter <= len(STEP_MARKS) else f"({step_counter})"

                # B列: ステップマーク（①②③...）
                ws.cell(row=cur_row, column=2, value=mark).font = font9

                # C列: サブ番号
                ws.cell(row=cur_row, column=3, value=sub_counter).font = font9
            else:
                sub_counter += 1
                # C列: サブ番号のみ
                ws.cell(row=cur_row, column=3, value=sub_counter).font = font9

            # D列: 操作種別
            ws.cell(row=cur_row, column=4, value=op_type).font = font9

            # H列: 対象データ（使用データ）
            if used_data:
                used_cell = ws.cell(row=cur_row, column=8, value=used_data)
                used_cell.font = font9
                used_cell.alignment = wrap

            # I列: 加工内容（設定値）
            if settings:
                settings_cell = ws.cell(row=cur_row, column=9, value=settings)
                settings_cell.font = font9
                settings_cell.alignment = wrap

            cur_row += 1

    cur_row += 1  # 空行

    # ========== ■ 最終確認 セクション ==========
    if check_sec and check_sec.get("rows"):
        write_section_header(cur_row, "■ 最終確認")
        cur_row += 1

        # ヘッダー行
        for j, col_name in enumerate(check_sec.get("columns", [])):
            cell = ws.cell(row=cur_row, column=j + 2, value=col_name)
            cell.font = font9_bold
            cell.fill = section_fill
        cur_row += 1

        for check_row in check_sec["rows"]:
            for c_idx, val in enumerate(check_row):
                cell = ws.cell(row=cur_row, column=c_idx + 2, value=str(val) if val else "")
                cell.font = font9
                cell.alignment = wrap
            cur_row += 1

    cur_row += 1

    # ========== ■ 検証観点 セクション ==========
    if verify_sec and verify_sec.get("rows"):
        write_section_header(cur_row, "■ 検証観点")
        cur_row += 1

        for j, col_name in enumerate(verify_sec.get("columns", [])):
            cell = ws.cell(row=cur_row, column=j + 2, value=col_name)
            cell.font = font9_bold
            cell.fill = section_fill
        cur_row += 1

        for v_row in verify_sec["rows"]:
            for c_idx, val in enumerate(v_row):
                cell = ws.cell(row=cur_row, column=c_idx + 2, value=str(val) if val else "")
                cell.font = font9
                cell.alignment = wrap
            cur_row += 1

    # --- 印刷設定 ---
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    title = generation_data.get("title", "データパレット構築手順書")
    filename = f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    return str(filepath), filename


# --- APIエンドポイント ---

@app.post("/api/upload-design")
async def upload_design(file: UploadFile = File(...)):
    """設計書JSONをアップロード"""
    suffix = Path(file.filename).suffix.lower()
    if suffix != ".json":
        raise HTTPException(400, "JSON形式のファイルをアップロードしてください")

    content = await file.read()
    try:
        design_doc = json.loads(content.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        raise HTTPException(400, f"JSONの解析に失敗しました: {e}")

    # バリデーション
    if not isinstance(design_doc, dict):
        raise HTTPException(400, "設計書の形式が不正です")

    return {"status": "ok", "design_document": design_doc, "filename": file.filename}


@app.post("/api/generate", response_model=GenerateResponse)
async def generate(req: GenerateRequest):
    """設計書から手順書を一発生成"""
    session_id = req.session_id or str(uuid.uuid4())

    design_text = format_design_document(req.design_document)
    system_prompt = SYSTEM_PROMPT_PHASE2.format(design_document=design_text)

    user_message = (
        "設計書の内容に基づいて、完全な手順書を生成してください。\n"
        "すべてのステップを省略なく、具体的に記述してください。\n"
        "質問は不要です。設計書にすべての情報が含まれています。"
    )

    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=16384,
            system=system_prompt,
            messages=[{"role": "user", "content": user_message}],
        )
        assistant_text = response.content[0].text
    except anthropic.APIError as e:
        return GenerateResponse(
            session_id=session_id,
            status="error",
            message=f"API呼び出しエラー: {e}",
        )

    # JSON抽出
    if "```json" in assistant_text:
        try:
            json_str = assistant_text.split("```json")[1].split("```")[0].strip()
            generation_data = json.loads(json_str)

            if generation_data.get("action") == "generate":
                filepath, filename = build_spreadsheet(generation_data)
                sessions[session_id] = {
                    "last_file": filepath,
                    "last_filename": filename,
                }

                return GenerateResponse(
                    session_id=session_id,
                    status="done",
                    message="手順書を生成しました！以下からダウンロードできます。",
                    download_url=f"/api/download/{session_id}",
                )
        except (json.JSONDecodeError, IndexError) as e:
            return GenerateResponse(
                session_id=session_id,
                status="error",
                message=f"生成結果の解析に失敗しました。もう一度お試しください。\nエラー: {e}",
            )

    return GenerateResponse(
        session_id=session_id,
        status="error",
        message=f"手順書の生成に失敗しました。AIからの応答:\n{assistant_text[:500]}",
    )


@app.get("/api/download/{session_id}")
async def download(session_id: str):
    session = sessions.get(session_id)
    if not session or "last_file" not in session:
        raise HTTPException(404, "ファイルが見つかりません")
    return FileResponse(
        session["last_file"],
        filename=session["last_filename"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# --- フロントエンド配信 ---
FRONTEND_PATH = BASE_DIR / "frontend"

@app.get("/", response_class=HTMLResponse)
async def index():
    return (FRONTEND_PATH / "phase2.html").read_text(encoding="utf-8")

@app.get("/doala.png")
async def doala_image():
    return FileResponse(FRONTEND_PATH / "doala.png", media_type="image/png")
