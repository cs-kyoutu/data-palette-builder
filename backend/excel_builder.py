"""Excel手順書ビルダー"""
import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

def build_spreadsheet(generation_data: dict) -> tuple[str, str]:
    """生成データから1シート形式のExcelファイルを作成（参考スプレッドシート準拠）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "手順書"

    # --- スタイル定義 ---
    GRAY_BG = "EFEFEF"
    font9 = Font(name="Yu Gothic UI", size=9)
    font9_bold = Font(name="Yu Gothic UI", size=9, bold=True)
    font10 = Font(name="Yu Gothic UI", size=10)
    section_fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    STEP_MARKS = ["①", "②", "③", "④", "⑤", "⑥", "⑦", "⑧", "⑨", "⑩",
                  "⑪", "⑫", "⑬", "⑭", "⑮", "⑯", "⑰", "⑱", "⑲", "⑳"]

    # --- 列幅設定（出力列: A〜F） ---
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 8    # 対象作業No（①②③）
    ws.column_dimensions["C"].width = 8    # 作業詳細No.
    ws.column_dimensions["D"].width = 14   # アイコン（操作名）
    ws.column_dimensions["E"].width = 25   # 作成後項目名
    ws.column_dimensions["F"].width = 70   # 手順書（完成形テキスト）

    # --- セクション帯を書く関数 ---
    def write_section_header(row, title_text):
        for col in range(1, 21):  # A-T
            cell = ws.cell(row=row, column=col)
            cell.fill = section_fill
        ws.cell(row=row, column=2, value=title_text).font = font9_bold

    # --- セクション探索 ---
    sections = generation_data.get("sections", [])
    overview_sec = next((s for s in sections if s.get("sheet_name") == "概要"), None)
    prep_sec = next((s for s in sections if "準備" in s.get("sheet_name", "")), None)
    proc_sec = next((s for s in sections if "加工" in s.get("sheet_name", "") or "結合" in s.get("sheet_name", "")), None)
    check_sec = next((s for s in sections if "確認" in s.get("sheet_name", "")), None)
    verify_sec = next((s for s in sections if "検証" in s.get("sheet_name", "")), None)

    cur_row = 1

    # ========== ■ フロー セクション ==========
    write_section_header(cur_row, "■ フロー")
    cur_row += 1
    cur_row += 1  # 空行

    # 横型フロー図: [Input1][Input2] → ①Op → ②Op → [Output]
    input_names = generation_data.get("input_tables", [])
    output_name = generation_data.get("output_name", "")
    steps_for_flow = []
    if proc_sec and proc_sec.get("rows"):
        marks = ["①","②","③","④","⑤","⑥","⑦","⑧","⑨","⑩","⑪","⑫","⑬","⑭","⑮"]
        mark_idx = 0
        for r in proc_sec["rows"]:
            if len(r) > 1 and r[0]:
                op_name = r[1] if len(r) > 1 else ""
                save_as = r[4] if len(r) > 4 else ""
                mark = marks[mark_idx] if mark_idx < len(marks) else f"({mark_idx+1})"
                steps_for_flow.append((mark, op_name, save_as))
                mark_idx += 1

    if input_names or steps_for_flow:
        thin = Side(style="thin")
        box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        BOX_W, BOX_H = 3, 2

        n_inputs = max(len(input_names), 1)
        # フロー全体の縦の中心行（インプット群の中心）
        total_input_height = n_inputs * BOX_H + (n_inputs - 1)  # BOX_H行 + 1行ギャップ
        mid_row = cur_row + total_input_height // 2

        def draw_box(r, c, text, fill_color=None):
            try:
                ws.merge_cells(start_row=r, start_column=c, end_row=r + BOX_H - 1, end_column=c + BOX_W - 1)
            except Exception:
                pass
            cell = ws.cell(row=r, column=c, value=text)
            cell.font = font9
            cell.alignment = center_align
            for rr in range(r, r + BOX_H):
                for cc in range(c, c + BOX_W):
                    ws.cell(row=rr, column=cc).border = box_border
            if fill_color:
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                for rr in range(r, r + BOX_H):
                    for cc in range(c, c + BOX_W):
                        ws.cell(row=rr, column=cc).fill = fill

        def draw_arrow(r, c):
            cell = ws.cell(row=r, column=c, value="→")
            cell.font = font9
            cell.alignment = center_align

        col = 2  # B列スタート

        # インプットボックス（左側、縦積み）
        for i, name in enumerate(input_names[:8]):
            r = cur_row + i * (BOX_H + 1)
            draw_box(r, col, name, fill_color="D9EAF7")

        col += BOX_W

        # インプット→ステップ1の矢印
        draw_arrow(mid_row, col)
        col += 1

        # ステップボックス（横並び）
        for mark, op_name, save_as in steps_for_flow[:10]:
            draw_box(mid_row - BOX_H // 2 + 1, col, f"{mark}\n{op_name}", fill_color="FFF9C4")
            col += BOX_W
            draw_arrow(mid_row, col)
            col += 1

        # アウトプットボックス
        if output_name:
            draw_box(mid_row - BOX_H // 2 + 1, col, output_name, fill_color="D5F5E3")

        cur_row += total_input_height + 2

    cur_row += 1  # 空行

    # ========== ■ 手順書 セクション ==========
    write_section_header(cur_row, "■ 手順書")
    cur_row += 1

    # ヘッダー行（出力列: B, C, D, E, F のみ）
    headers = ["", "対象\n作業No", "作業\n詳細No.", "アイコン", "作成後\n項目名", "手順書"]
    for j, h in enumerate(headers):
        cell = ws.cell(row=cur_row, column=j + 1, value=h)
        cell.font = font9_bold
        cell.alignment = wrap
    cur_row += 1

    if proc_sec and proc_sec.get("rows"):
        proc_rows = proc_sec["rows"]
        step_counter = 0

        for p_idx, p_row in enumerate(proc_rows):
            step_val = p_row[0] if len(p_row) > 0 else ""
            op_type = p_row[1] if len(p_row) > 1 else ""
            save_as = p_row[4] if len(p_row) > 4 else ""
            complete_text = p_row[11] if len(p_row) > 11 else ""

            # B列: 対象作業No（①②③）
            if step_val and str(step_val).strip():
                mark = STEP_MARKS[step_counter] if step_counter < len(STEP_MARKS) else f"({step_counter+1})"
                ws.cell(row=cur_row, column=2, value=mark).font = font9
                step_counter += 1

            # C列: 作業詳細No.
            sub_num = str(step_val).strip() if step_val else ""
            if sub_num:
                try:
                    ws.cell(row=cur_row, column=3, value=float(sub_num)).font = font9
                except (ValueError, TypeError):
                    ws.cell(row=cur_row, column=3, value=sub_num).font = font9

            # D列: アイコン（操作名）
            ws.cell(row=cur_row, column=4, value=op_type).font = font9

            # E列: 作成後項目名
            if save_as:
                ws.cell(row=cur_row, column=5, value=save_as).font = font9

            # F列: 手順書（完成形テキスト）
            if complete_text:
                comp_cell = ws.cell(row=cur_row, column=6, value=complete_text)
                comp_cell.font = font9
                comp_cell.alignment = wrap

            cur_row += 1

    cur_row += 2  # 空行

    # ========== ■ 最終確認 セクション ==========
    if check_sec and check_sec.get("rows"):
        write_section_header(cur_row, "■ 最終確認")
        cur_row += 1

        # ヘッダー行
        for j, col_name in enumerate(check_sec.get("columns", [])):
            cell = ws.cell(row=cur_row, column=j + 2, value=col_name)
            cell.font = font9_bold
            cell.fill = section_fill
        cur_row += 1

        for check_row in check_sec["rows"]:
            for c_idx, val in enumerate(check_row):
                cell = ws.cell(row=cur_row, column=c_idx + 2, value=str(val) if val else "")
                cell.font = font9
                cell.alignment = wrap
            cur_row += 1

    cur_row += 1

    # ========== ■ 検証観点 セクション ==========
    if verify_sec and verify_sec.get("rows"):
        write_section_header(cur_row, "■ 検証観点")
        cur_row += 1

        for j, col_name in enumerate(verify_sec.get("columns", [])):
            cell = ws.cell(row=cur_row, column=j + 2, value=col_name)
            cell.font = font9_bold
            cell.fill = section_fill
        cur_row += 1

        for v_row in verify_sec["rows"]:
            for c_idx, val in enumerate(v_row):
                cell = ws.cell(row=cur_row, column=c_idx + 2, value=str(val) if val else "")
                cell.font = font9
                cell.alignment = wrap
            cur_row += 1

    # --- 印刷設定 ---
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    title = generation_data.get("title", "データパレット構築手順書")
    # ファイル名に使えない文字をサニタイズ
    import re
    safe_title = re.sub(r'[/\\:*?"<>|]', '_', title)[:50]
    filename = f"{safe_title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    return str(filepath), filename


# --- APIエンドポイント ---

